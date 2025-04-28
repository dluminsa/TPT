import pandas as pd 
import streamlit as st 
import os
import numpy as np
import gspread
from openpyxl import load_workbook
from pathlib import Path
import traceback
import time
from google.oauth2.service_account import Credentials
from oauth2client.service_account import ServiceAccountCredentials
from streamlit_gsheets import GSheetsConnection
from datetime import datetime 
import datetime as dt
#st.stop()
#Clear cache at the very start of the app
st.cache_data.clear()
st.cache_resource.clear()

#sddd

# st.write('CLOSED, RETURNING NEXT QUER')
# st.write('**MERRY XMASS**')
# time.sleep(4)
# st.balloons()
# st.write('**AND A HAPPY NEW YEAR**')
# a = [1,2,3,4]
# for i in a:
#     st.balloons()
#     time.sleep(4)         
# st.stop()

cola,colb,colc = st.columns([1,3,1])
colb.subheader('AUTOMATICALLY GENERATE TPT LINELISTS')

today = datetime.now()
todayd = today.strftime("%Y-%m-%d")# %H:%M")
week = today.strftime("%V")
wk = int(week) + 13
cola,colb = st.columns(2)
cola.write(f"**DATE TODAY:    {todayd}**")
colb.write(f"**CURRENT WEEK:    {wk}**")
st.image('rename.png', caption='instructions')
# st.image("BEFORE.png", caption="BEFORE")
# st.image("AFTER.png", caption="AFTER")


# Display the HTML table using markdown in Streamlit
#st.markdown(html_table, unsafe_allow_html=True)
st.markdown('**AFTER, SAVE THE EXTRACT AS an XLSX BEFORE YOU PROCEED**')

file = st.file_uploader("Upload your EMR extract here", type=['xlsx']) 

if file is not None:   
    if 'fd' not in st.session_state:
        fileN = file.name
        name = os.path.basename(fileN).split('.')[0]
        st.session_state.fd = name
    else:
        pass
else:
    pass
if file is not None: 
   fileN = file.name
   namey = os.path.basename(fileN).split('.')[0]
   if str(namey) != str(st.session_state.fd):
            #st.info(f'DATA FOR {facy} NOT SUBMITTED')
            st.session_state.submited = False
            st.cache_data.clear()
            st.session_state.fd = namey
            st.cache_resource.clear()
            st.session_state.submited =False
            st.session_state.df = None
            st.session_state.reader =False#
            time.sleep(1)
            st.rerun()
        
if 'submited' not in st.session_state:
    st.session_state.submited =False
if 'df' not in st.session_state:
    st.session_state.df = None
# if 'fac' not in st.session_state:
# st.session_state.fac = None
if 'reader' not in st.session_state:
    st.session_state.reader =False#
#ext = None
if file is not None and not st.session_state.reader:
    # Get the file name
    fileN = file.name
    ext = os.path.basename(fileN).split('.')[1]
    if ext == 'xlsx.xlsx':
           ext = 'xlsx'
#df = None
if file is not None and not st.session_state.reader:
    wb = load_workbook(file)
    sheets = wb.sheetnames
    if len(sheets)>1:
        st. warning('THIS EXTRACT HAS MULTIPLE SHEETS, I CAN NOT TELL WHICH ONE TO READ')
        time.sleep(3)
        st.info('DELETE ALL THE OTHER SHEETS AND REMAIN WITH ONE THAT HAS THE EVER ENROLLED')
        st.stop()
    else:
        pass
    if ext !='xlsx':
        pass
    else:
                st.session_state.df = pd.read_excel(file)
                df = st.session_state.df
                st.write('Excel accepted, summaries and linelists below will be for this excel')
                st.write('To change this excel or to upload another excel, first refresh the page')
#if file is not None and not st.session_state.reader:
                df = df.rename(columns= {'ART  ':'ART',  'AS  ':'AS', 'RD  ':'RD',  'LD  ': 'LD', 'TO  ':'TO', 'DD  ': 'DD', 'TPT  ':'TPT'})
                df = df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'RD ':'RD',  'LD ': 'LD', 'TO ':'TO', 'DD ': 'DD',  'TPT ':'TPT'})
                columns = ['ART','AS', 'RD','TO', 'DD','LD','TPT']
                cols = df.columns.to_list()
                if not all(column in cols for column in columns):
                    missing_columns = [column for column in columns if column not in cols]
                    for column in missing_columns:
                        st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                        st.markdown('**First rename all the columns as guided above**')
                        st.stop()
                st.session_state.reader= True
if st.session_state.reader:
                      # Convert 'ART' column to string and create 'ART' column with numeric part to remove blanks
                    st.session_state.df = st.session_state.df.rename(columns= {'ART  ':'ART', 'TPT  ':'TPT', 'AS  ':'AS', 'RD  ':'RD', 'LD  ': 'LD',  'TO  ':'TO', 'DD  ': 'DD'})
                    st.session_state.df = st.session_state.df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'TPT ':'TPT','RD ':'RD', 'RDO ':'RDO', 'RD1 ':'RD1', 'RD2 ':'RD2', 'VD ':'VD', 'FE ':'FE', 'LD ': 'LD', 'ARVD ': 'ARVD',
                        'TO ':'TO', 'DD ': 'DD','TPT ':'TPT'})
    
                    df = st.session_state.df[['ART','AS', 'RD', 'TO', 'DD','LD', 'TPT']].copy()
                    df['ART'] = df['ART'].astype(str)
                    df['A'] = df['ART'].str.replace('[^0-9]', '', regex=True)
                    df['A'] = pd.to_numeric(df['A'], errors= 'coerce')
                    df = df[df['A']>0].copy()
                    #df.dropna(subset='ART', inplace=True)
                    
                    df[['AS', 'RD','TO', 'DD', 'LD']] = df[['AS', 'RD','TO', 'DD', 'LD']].astype(str)
                    if df['TO'].str.contains('YES').any():
                        st.write("You may be using the Transfer in column instead of the Transfer_in Obs date column")
                        st.stop()
                    
                    df['AS'] = df['AS'].astype(str)
                    df['RD'] = df['RD'].astype(str)
                    df['TO'] = df['TO'].astype(str)
                    
                    df['DD'] = df['DD'].astype(str)
                    df['LD'] = df['LD'].astype(str)
                    
                    
                    y = pd.DataFrame({'ART' :['2','3','4','5'], 'TI':['1-1-1',1,'1/1/1','3 8 2001'], 'RD':['1-1-1',1,'1/1/1','3 8 2001'],'DD':['1-1-1',1,'1/1/1','3 8 2001'], 
                                    'TO':['1-1-1',1,'1/1/1','3 8 2001'], 'AS':['1-1-1',1,'1/1/1','3 8 2001'],'LD':['1-1-1',1,'1/1/1','3 8 2001']})   
    
                    
                    df['AS'] = df['AS'].astype(str)
                    df['RD'] = df['RD'].astype(str)
                    df['TI'] = df['TI'].astype(str)
                    df['TO'] = df['TO'].astype(str)
                    df['DD'] = df['DD'].astype(str)
                    df['LD'] = df['LD'].astype(str)
                  
        
                    df['AS'] = df['AS'].str.replace('00:00:00', '', regex=True)
                    df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
             
                   
                    df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
                    
                    df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
                    df['LD'] = df['LD'].str.replace('00:00:00', '', regex=True)
                    
        
                    df = pd.concat([df,y])
                    df['AS'] = df['AS'].astype(str)
                    df['RD'] = df['RD'].astype(str)
                  
                    df['TO'] = df['TO'].astype(str)
                    df['DD'] = df['DD'].astype(str)
                    df['LD'] = df['LD'].astype(str)        
      
        
        
                    # SPLITTING ART START DATE
                    A = df[df['AS'].str.contains('-')].copy()
                    a = df[~df['AS'].str.contains('-')].copy()
                    B = a[a['AS'].str.contains('/')].copy()
                    C = a[~a['AS'].str.contains('/')].copy()
                    E = C[C['AS'].str.contains(' ')].copy()
                    D = C[~C['AS'].str.contains(' ')].copy()
        
                    A[['Ayear', 'Amonth', 'Aday']] = A['AS'].str.split('-', expand = True)
                    B[['Ayear', 'Amonth', 'Aday']] = B['AS'].str.split('/', expand = True)
                    try:
                        D['AS'] = pd.to_numeric(D['AS'], errors='coerce')
                        D['AS'] = pd.to_datetime(D['AS'], origin='1899-12-30', unit='D', errors='coerce')
                        D['AS'] =  D['AS'].astype(str)
                        D[['Ayear', 'Amonth', 'Aday']] = D['AS'].str.split('-', expand = True)
                    except:
                        pass
                    try:  
                        E['AS'] = pd.to_datetime(E['AS'],format='%d %m %Y', errors='coerce')
                        E['AS'] =  E['AS'].astype(str)
                        E[['Ayear', 'Amonth', 'Aday']] = E['AS'].str.split('-', expand = True)
                    except:
                        pass
                    df = pd.concat([A,B,D,E]) 
        
                    # SPLITTING DEATH DATE
                    A = df[df['DD'].str.contains('-')].copy()
                    a = df[~df['DD'].str.contains('-')].copy()
                    B = a[a['DD'].str.contains('/')].copy()
                    C = a[~a['DD'].str.contains('/')].copy()
                    E = C[C['DD'].str.contains(' ')].copy()
                    D = C[~C['DD'].str.contains(' ')].copy()
                    A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
                    B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)

                    try:
                        D['DD'] = pd.to_numeric(D['DD'], errors='coerce')
                        D['DD'] = pd.to_datetime(D['DD'], origin='1899-12-30', unit='D', errors='coerce')
                        D['DD'] =  D['DD'].astype(str)
                        D[['Dyear', 'Dmonth', 'Dday']] = D['DD'].str.split('-', expand = True)
                    except:
                        pass
                    try:  
                        E['DD'] = pd.to_datetime(E['DD'],format='%d %m %Y', errors='coerce')
                        E['DD'] =  E['DD'].astype(str)
                        E[['Dyear', 'Dmonth', 'Dday']] = E['DD'].str.split('-', expand = True)
                    except:
                        pass
                    df = pd.concat([A,B,D,E]) 
                
                    # SORTING THE RETURN VISIT DATE
                    A = df[df['RD'].str.contains('-')].copy()
                    a = df[~df['RD'].str.contains('-')].copy()
                    B = a[a['RD'].str.contains('/')].copy()
                    C = a[~a['RD'].str.contains('/')].copy()
                    E = C[C['RD'].str.contains(' ')].copy()
                    D = C[~C['RD'].str.contains(' ')].copy()
                   
                    #D = C[C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
                    #E = C[~C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
            
                    A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
                    B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
                    try:
                        D['RD'] = pd.to_numeric(D['RD'], errors='coerce')
                        D['RD'] = pd.to_datetime(D['RD'], origin='1899-12-30', unit='D', errors='coerce')
                        D['RD'] =  D['RD'].astype(str)
                        D[['Ryear', 'Rmonth', 'Rday']] = D['RD'].str.split('-', expand = True)
                    except:
                        pass
                    try:  
                        E['RD'] = pd.to_datetime(E['RD'],format='%d %m %Y', errors='coerce')
                        E['RD'] =  E['RD'].astype(str)
                        E[['Ryear', 'Rmonth', 'Rday']] = E['RD'].str.split('-', expand = True)
                    except:
                        pass
                    df = pd.concat([A,B,D,E]) 
                
                   
                    #SORTING THE TO DATE
                    A = df[df['TO'].str.contains('-')].copy()
                    a = df[~df['TO'].str.contains('-')].copy()
                    B = a[a['TO'].str.contains('/')].copy()
                    C = a[~a['TO'].str.contains('/')].copy()
                    E = C[C['TO'].str.contains(' ')].copy()
                    D = C[~C['TO'].str.contains(' ')].copy()
        
                    A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
                    B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
                    try:
                        D['TO'] = pd.to_numeric(D['TO'], errors='coerce')
                        D['TO'] = pd.to_datetime(D['TO'], origin='1899-12-30', unit='D', errors='coerce')
                        D['TO'] =  D['TO'].astype(str)
                        D[['Tyear', 'Tmonth', 'Tday']] = D['TO'].str.split('-', expand = True)
                    except:
                        pass
                    try:  
                        E['TO'] = pd.to_datetime(E['TO'],format='%d %m %Y', errors='coerce')
                        E['TO'] =  E['TO'].astype(str)
                        E[['Tyear', 'Tmonth', 'Tday']] = E['TO'].str.split('-', expand = True)
                    except:
                        pass
                    df = pd.concat([A,B,D,E])
        
                    # SORTING THE LAST ENCOUNTER DATES
                    A = df[df['LD'].str.contains('-')].copy()
                    a = df[~df['LD'].str.contains('-')].copy()
                    B = a[a['LD'].str.contains('/')].copy()
                    C = a[~a['LD'].str.contains('/')].copy()
                    E = C[C['LD'].str.contains(' ')].copy()
                    D = C[~C['LD'].str.contains(' ')].copy()
            
                    A[['Lyear', 'Lmonth', 'Lday']] = A['LD'].str.split('-', expand = True)
                    B[['Lyear', 'Lmonth', 'Lday']] = B['LD'].str.split('/', expand = True)
                    try:
                        D['LD'] = pd.to_numeric(D['LD'], errors='coerce')
                        D['LD'] = pd.to_datetime(D['LD'], origin='1899-12-30', unit='D', errors='coerce')
                        D['LD'] =  D['LD'].astype(str)
                        D[['Lyear', 'Lmonth', 'Lday']] = D['LD'].str.split('-', expand = True)
                    except:
                        pass
                    try:  
                        E['LD'] = pd.to_datetime(E['LD'],format='%d %m %Y', errors='coerce')
                        E['LD'] =  E['LD'].astype(str)
                        E[['Lyear', 'Lmonth', 'Lday']] = E['LD'].str.split('-', expand = True)
                    except:
                        pass
                    df = pd.concat([A,B,D,E])
                
                    
        
                    #BRINGING BACK THE / IN DATES
                    df['AS'] = df['AS'].astype(str)
                    df['ARVDO'] = df['ARVDO'].astype(str)
                    df['RD'] = df['RD'].astype(str)
                    df['TO'] = df['TO'].astype(str)
                    
                    df['DD'] = df['DD'].astype(str)
                    df['LD'] = df['LD'].astype(str)
             
        #             #Clearing NaT from te dates
                    df['AS'] = df['AS'].str.replace('NaT', '',regex=True)
                    df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
                  
                    df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
                    
                    df['DD'] = df['DD'].str.replace('NaT', '',regex=True)
                    df['LD'] = df['LD'].str.replace('NaT', '',regex=True)
                    
        
                    # #SORTING THE RETURN VISIT DATE YEARS
                    df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                    
                    df['Ryear'] = df['Ryear'].fillna(994)
                    a = df[df['Ryear']>31].copy()
                    b = df[df['Ryear']<32].copy()
                    b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
                    b = b.rename(columns={'Rday2': 'Rday'})
        
                    df = pd.concat([a,b])
                    dfc = df.shape[0]
                    
                        #SORTING THE TRANSFER OUT DATE YEAR
                    df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
                    df['Tyear'] = df['Tyear'].fillna(994)
                    a = df[df['Tyear']>31].copy()
                    b = df[df['Tyear']<32].copy()
                    b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
                    b = b.rename(columns={'Tday2': 'Tday'})
                    df = pd.concat([a,b])         
                    
                    #SORTING THE ART START YEARS
                    df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
                    df['Ayear'] = df['Ayear'].fillna(994)
                    a = df[df['Ayear']>31].copy()
                    b = df[df['Ayear']<32].copy()
                    b = b.rename(columns={'Ayear': 'Aday2', 'Aday': 'Ayear'})
                    b = b.rename(columns={'Aday2': 'Aday'})
                    df = pd.concat([a,b])
                    dfe = df.shape[0]
        
                    #SORTING THE ART START YEARS
                    df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
                    df['Dyear'] = df['Dyear'].fillna(994)
                    a = df[df['Dyear']>31].copy()
                    b = df[df['Dyear']<32].copy()
                    b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
                    b = b.rename(columns={'Dday2': 'Dday'})
                    df = pd.concat([a,b])
                    dfe = df.shape[0]
        
                    # #SORTING THE LAST ENCOUNTER
                    df[['Lday', 'Lyear']] = df[['Lday', 'Lyear']].apply(pd.to_numeric, errors='coerce')
                    
                    df['Lyear'] = df['Lyear'].fillna(994)
                    a = df[df['Lyear']>31].copy()
                    b = df[df['Lyear']<32].copy()
                    b = b.rename(columns={'Lyear': 'Lday2', 'Lday': 'Lyear'})
                    b = b.rename(columns={'Lday2': 'Lday'})
        
                    df = pd.concat([a,b])
                    dfc = df.shape[0]
        
                    
        
                    #CREATE WEEKS 
                    df['Rdaya'] = df['Rday'].astype(str).str.split('.').str[0]
                    df['Rmontha'] = df['Rmonth'].astype(str).str.split('.').str[0]
                    df['Ryeara'] = df['Ryear'].astype(str).str.split('.').str[0]
        
                    df['RETURN DATE'] = df['Rdaya'] + '/' + df['Rmontha'] + '/' + df['Ryeara']
                    df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
                    #CREATING WEEEK FOR RETURN VISIT DATE
                    df['RWEEK'] = df['RETURN DATE'].dt.strftime('%V')
                    df['RWEEK'] = pd.to_numeric(df['RWEEK'], errors='coerce')
                    df['RWEEK1'] = df['RWEEK'] + 13
    
                    today = dt.date.today() 
                    todayr = pd.to_datetime(today)
                    #df['DURL'] = round((todayr)-(df['RETURN DATE']))#.dt.days / 30)  
                    df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'])
    
    
                    #POTENTIAL TXCUR ALTER... 
                    df[['Rmonth', 'Rday', 'Ryear']] = df[['Rmonth', 'Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                    df25 = df[df['Ryear']>2024].copy()
                    df24 = df[df['Ryear'] == 2024].copy()
                    df24[['Rmonth', 'Rday']] = df24[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                    df24 = df24[((df24['Rmonth']>12) | ((df24['Rmonth']==12) & (df24['Rday']>3)))].copy()
                    df = pd.concat([df25, df24]).copy()
                    pot = df.shape[0]
        
                    #REMOVE TO of the last reporting month
                    df[ 'Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
                    df = df[df['Tyear']==994].copy()
                    
    
                    #REMOVE the dead of the reporting month
                    df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
                    df = df[df['Dyear']==994].copy()
                        
                        
if st.session_state.reader:                                                    
    file2 = r'CLUSTERS.csv'
    dfx = pd.read_csv(file2)
    clusters  = list(dfx['CLUSTER'].unique())
    cluster = st.radio(label='**Choose your cluster**', options=clusters,index=None, horizontal=True)
    if not cluster:
        st.stop()
    else:
        districts = dfx[dfx['CLUSTER']==cluster]
        districts = list(districts['DISTRICT'].unique())
        district = st.radio(label='**Choose your district**', options=districts,index=None, horizontal=True)
        if not district:
            st.stop()
        else:
            facilities = dfx[dfx['DISTRICT']==district]
            facilities = facilities['FACILITY'].unique()
            facility = st.selectbox(label='**Choose your facility**', options=facilities,index=None)
            if not facility:
                st.stop()
            else:
                facy = facility
                #st.session_state.fac = facility
                if 'fac' not in st.session_state:
                       st.session_state.fac = facility
                pass
                if str(facy) != str(st.session_state.fac):
                        #st.info(f'DATA FOR {facy} NOT SUBMITTED')
                        st.session_state.submited = False
                        st.session_state.fac = facy
                        st.cache_data.clear()
                        st.cache_resource.clear()
                        st.session_state.submited =False
                        st.session_state.df = None
                        st.session_state.reader =False#
                        time.sleep(1)
                        st.rerun()
    
if st.session_state.reader:# and st.session_state.df:
                @st.cache_data
                def yearto6():
                    dat = newto6.copy()
                    dat = dat[['ART','AS', 'RD', 'VD']]
                    dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                    return dat

                row7 = part +list7
                    
                secrets = st.secrets["connections"]["gsheets"]
                
                    # Prepare the credentials dictionary
                credentials_info = {
                        "type": secrets["type"],
                        "project_id": secrets["project_id"],
                        "private_key_id": secrets["private_key_id"],
                        "private_key": secrets["private_key"],
                        "client_email": secrets["client_email"],
                        "client_id": secrets["client_id"],
                        "auth_uri": secrets["auth_uri"],
                        "token_uri": secrets["token_uri"],
                        "auth_provider_x509_cert_url": secrets["auth_provider_x509_cert_url"],
                        "client_x509_cert_url": secrets["client_x509_cert_url"]
                    }
                        
                try:
                    # Define the scopes needed for your application
                    scopes = ["https://www.googleapis.com/auth/spreadsheets",
                            "https://www.googleapis.com/auth/drive"]
                    
                     
                    credentials = Credentials.from_service_account_info(credentials_info, scopes=scopes)
                        
                        # Authorize and access Google Sheets
                    client = gspread.authorize(credentials)
                        
                        # Open the Google Sheet by URL
                    spreadsheetu = "https://docs.google.com/spreadsheets/d/1twNlv9MNQWWsM73_dA19juHkp_Hua-k-fJA1qNVwQl0"
                    spreadsheet = client.open_by_url(spreadsheetu)
                except Exception as e:
                        # Log the error message
                    st.write(f"CHECK: {e}")
                    st.write(traceback.format_exc())
                    st.write("COULDN'T CONNECT TO GOOGLE SHEET, TRY AGAIN")
                    st.stop()

                #LINE LISTS         
                df[['Ryear', 'Rmonth', 'Rday']] = df[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                line = df[((df['Ryear'] == 2025) & (df['Rmonth'].isin([4,5,6])))].copy()
                tpt = line.copy()
    
                tpta  = tpt[tpt['TPT'].notna()].copy()
                tptb  = tpt[tpt['TPT'].isnull()].copy()
                tpta['TPT'] = tpta['TPT'].astype(str)
                tpta = tpta[tpta['TPT']=='Never'].copy()
    
                tpt = pd.concat([tpta, tptb]) #NEVER AND BLANKS
                month = dt.date.today().strftime('%m')
                mon = int(month)
    
                tpt[['Ayear', 'Amonth']] = tpt[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                tpta = tpt[((tpt['Ayear'] ==2025) & (tpt['Amonth'].isin([1,2,3])))].copy()
    
                tptb = tpt[((tpt['Ayear'] <2025)| ((tpt['Ayear'] ==2025) & (tpt['Amonth']<4)))].copy() #NEXT Q ALL 2024 WILL BE ELIGIBLE
    
                tpta[['Ayear', 'Rmonth']] = tpta[['Ayear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                tpta['CHECK'] = tpt['Rmonth']- tpt['Amonth'].copy()
                tpta['CHECK'] = pd.to_numeric(tpta['CHECK'], errors = 'coerce')
                tpta = tpta[tpta['CHECK']>2].copy()
                tpt = pd.concat([tpta, tptb])
    
                #likely Vs unlikely
                tpt[['Ayear', 'Amonth']] = tpt[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                tpta = tpt[((tpt['Ayear']<2024) | ((tpt['Ayear']==2024) & (tpt['Amonth'] <7)))].copy()
                tptb = tpt[((tpt['Ayear']==2024) & (tpt['Amonth'] >6))].copy()
                tpta['TPT STATUS'] = 'UNLIKELY'
                tptb['TPT STATUS'] = 'LIKELY'
                tpt = pd.concat([tpta, tptb])
                tpt['Rmonth'] = pd.to_numeric(tpt['Rmonth'], errors = 'coerce')
                aprilpt = tpt[tpt['Rmonth']==4].shape[0]
                maytpt = tpt[tpt['Rmonth']==5].shape[0]
                junetpt = tpt[tpt['Rmonth']==6].shape[0]
                tpt = tpt[['A', 'TPT STATUS']] # GET RD,AS,RDAY,RMONTH, AFTER MERGING
                weeks = [15,16,17,18,19,20,21,22,23, 24,25,26]
                numb = []
                nom = []
                for wk in weeks:
                    tptx = tpt[tpt['RWEEK'] == wk].copy()
                    tpty = tptx[tptx['TPT STATUS'] == 'UNLIKELY'].copy()
                    nuf = tpty.shape[0]
                    tptx = tptx[tptx['TPT STATUS'] == 'LIKELY'].copy()
                    nub = tptx.shape[0]
                    numb.append(nub)
                    nom.append (nuf)
                    
                
                weekis = numb
                
                @st.cache_data
                def missedlists():
                    dat = tptcopy()
                    dat = dat.rename(columns={'LD': 'LAST ENCOUNTER', 'GD':'GENDER','AG':'AGE', 'RD':'RETURN DATE', 'A':'ART No.'})
                    dat = dat[['ART No.', 'RETURN DATE',  'LAST ENCOUNTER', 'TPT', 'TPT STATUS']].copy()
                    return dat

                #SUMMARY LINELIST
                col1,col2,col3 = st.columns([1,2,1])
                with col3:
                     submit = st.button('Submit') 

                linelists = [cluster, district, facility, jancx, janvl,jantpt, febcx, febvl, febtpt, marcx, marvl, martpt, notbled, notpt, notscreened, wk]
    
                if submit:
                        try:
                            sheet7 = spreadsheet.worksheet("Q4")
                            sheet7.append_row(row7, value_input_option='RAW')
                            sheet9 = spreadsheet.worksheet("ONEYR")
                            sheet9.append_row(list8, value_input_option='RAW')
                            st.session_state.submited = True
                        except Exception as e:
                            # Print the error message
                            st.write(f"ERROR: {e}")
                            st.stop()  # Stop the Streamlit app here to let the user manually retry     
                else:
                        st.write('FIRST SUBMIT TO SEE LINELISTS AND SUMMARY') 
                        st.markdown(f'**YOU HAVE SELECTED {district} AS THE DISTRICT AND {facility} AS THE FACILITY**')
                        st.write('BE SURE OF THE ABOVE SELECTIONS BEFORE SUBMITTING')                     
                
                if not st.session_state.submited:
                    st.stop()  
                if str(facy) != str(st.session_state.fac):
                    st.stop()
                if st.session_state.submited:
                        st.success('**SUBMITTED, To upload another excel, first refresh this page, or open the link afresh**')
                        st.divider()
                        if facility == 'Kifampa HC III':
                            pass
                        elif pot < prev:
                            st.info('**SOMETHING IS WRONG WITH THIS EXTRACT, SEND TO YOUR M AND E TO CHECK, POTENTIAL TX CURR IS LESS THAN Q1 CURR**')
                            st.stop()
                        else:
                            pass 
                        st.write(f"<h6><b>DOWNLOAD LINELISTS FROM HERE</b></h6>", unsafe_allow_html=True)
                        cola, colb, colc = st.columns(3)
                        with cola:
                                if lostq ==0:
                                    st.write('**NO TXML**')
                                else:
                                   dat = currlost.copy() 
                                   #dat = tttt.copy()
                                   csv_data = dat.to_csv(index=False)
                                   tot = dat.shape[0]
                                   st.write(f'**YOUR TXML IS {tot}**')
                                   st.download_button(
                                                label="Q2 TXML",
                                                data=csv_data,
                                                file_name=f"{facility} TXML.csv",
                                                mime="text/csv")
            
                        st.divider()
                        st.success('**WANT TO HELP US IMPROVE?**')
                        st.write('Are you getting different results when you filter the extract manually?, That is ok, You can send inyour extract for comparison')
                        st.write('')
                        st.write('')
                        st.write('')
                        st.success('**@ LUMINSA DESIRE**')
                        


