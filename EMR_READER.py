import pandas as pd 
import streamlit as st 
import os
import numpy as np
import gspread
from openpyxl import load_workbook
from pathlib import Path
import traceback
import time
from google.oauth2.service_account import Credentials
from oauth2client.service_account import ServiceAccountCredentials
from streamlit_gsheets import GSheetsConnection
from datetime import datetime 
import datetime as dt
#st.stop()
#Clear cache at the very start of the app
st.cache_data.clear()
st.cache_resource.clear()

#sddd

# st.write('CLOSED, RETURNING NEXT QUER')
# st.write('**MERRY XMASS**')
# time.sleep(4)
# st.balloons()
# st.write('**AND A HAPPY NEW YEAR**')
# a = [1,2,3,4]
# for i in a:
#     st.balloons()
#     time.sleep(4)         
# st.stop()
def extract():
    cola,colb,colc = st.columns([1,3,1])
    colb.subheader('PROGRAM GROWTH')
    
    today = datetime.now()
    todayd = today.strftime("%Y-%m-%d")# %H:%M")
    week = today.strftime("%V")
    wk = int(week) + 13
    cola,colb = st.columns(2)
    cola.write(f"**DATE TODAY:    {todayd}**")
    colb.write(f"**CURRENT WEEK:    {wk}**")
    st.image('rename.png', caption='instructions')
    st.image("BEFORE.png", caption="BEFORE")
    st.image("AFTER.png", caption="AFTER")
    
    
    # Display the HTML table using markdown in Streamlit
    #st.markdown(html_table, unsafe_allow_html=True)
    st.markdown('**AFTER, SAVE THE EXTRACT AS an XLSX BEFORE YOU PROCEED, Check User manual for further guidance**')
    
    file = st.file_uploader("Upload your EMR extract here", type=['xlsx']) 

    if file is not None:   
        if 'fd' not in st.session_state:
            fileN = file.name
            name = os.path.basename(fileN).split('.')[0]
            st.session_state.fd = name
        else:
            pass
    else:
        pass
    if file is not None: 
       fileN = file.name
       namey = os.path.basename(fileN).split('.')[0]
       if str(namey) != str(st.session_state.fd):
                #st.info(f'DATA FOR {facy} NOT SUBMITTED')
                st.session_state.submited = False
                st.cache_data.clear()
                st.session_state.fd = namey
                st.cache_resource.clear()
                st.session_state.submited =False
                st.session_state.df = None
                st.session_state.reader =False#
                time.sleep(1)
                st.rerun()
            
    if 'submited' not in st.session_state:
        st.session_state.submited =False
    if 'df' not in st.session_state:
        st.session_state.df = None
    # if 'fac' not in st.session_state:
   # st.session_state.fac = None
    if 'reader' not in st.session_state:
        st.session_state.reader =False#
    #ext = None
    if file is not None and not st.session_state.reader:
        # Get the file name
        fileN = file.name
        ext = os.path.basename(fileN).split('.')[1]
        if ext == 'xlsx.xlsx':
               ext = 'xlsx'
    #df = None
    if file is not None and not st.session_state.reader:
        wb = load_workbook(file)
        sheets = wb.sheetnames
        if len(sheets)>1:
            st. warning('THIS EXTRACT HAS MULTIPLE SHEETS, I CAN NOT TELL WHICH ONE TO READ')
            time.sleep(3)
            st.info('DELETE ALL THE OTHER SHEETS AND REMAIN WITH ONE THAT HAS THE EVER ENROLLED')
            st.stop()
        else:
            pass
        if ext !='xlsx':
            st.write('Unsupported file format, first save the excel as xlsx and try again')
            st.stop()
        else:
                    st.session_state.df = pd.read_excel(file)
                    df = st.session_state.df
                    st.write('Excel accepted, summaries and linelists below will be for this excel')
                    st.write('To change this excel or to upload another excel, first refresh the page')
    #if file is not None and not st.session_state.reader:
                    df = df.rename(columns= {'ART  ':'ART',  'AS  ':'AS', 'RD  ':'RD', 'RDO  ':'RDO', 'RD1  ':'RD1', 'RD2  ':'RD2', 'VD  ':'VD', 'FE  ':'FE', 'LD  ': 'LD', 'ARVD  ': 'ARVD',
       'ARVDO  ': 'ARVDO', 'TI  ': 'TI', 'TO  ':'TI', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD', 'PT  ': 'PT', 'TPT  ':'TPT', 'CX  ': 'CX'})
                    df = df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'RD ':'RD', 'RDO ':'RDO', 'RD1 ':'RD1', 'RD2 ':'RD2', 'VD ':'VD', 'FE ':'FE', 'LD ': 'LD', 'ARVD ': 'ARVD',
                           'ARVDO ': 'ARVDO', 'TI ': 'TI', 'TO ':'TI', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD', 'PT ': 'PT', 'TPT ':'TPT', 'CX ': 'CX'})
                    columns = ['ART','AG', 'GD','AS', 'VD', 'RD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'RDO', 'ARVD', 'ARVDO','TPT','CX', 'PT']
                    cols = df.columns.to_list()
                    if not all(column in cols for column in columns):
                        missing_columns = [column for column in columns if column not in cols]
                        for column in missing_columns:
                            st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                            st.markdown('**First rename all the columns as guided above**')
                            st.stop()
                    st.session_state.reader= True
    if st.session_state.reader:
                          # Convert 'ART' column to string and create 'ART' column with numeric part to remove blanks
                        st.session_state.df = st.session_state.df.rename(columns= {'ART  ':'ART', 'TPT  ':'TPT', 'AS  ':'AS', 'RD  ':'RD', 'RDO  ':'RDO', 'RD1  ':'RD1', 'RD2  ':'RD2', 'VD  ':'VD', 'FE  ':'FE', 'LD  ': 'LD', 'ARVD  ': 'ARVD',
       'ARVDO  ': 'ARVDO', 'TI  ': 'TI', 'TO  ':'TI', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD',  'CX  ': 'CX', 'PT  ': 'PT'})
                        st.session_state.df = st.session_state.df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'TPT ':'TPT','RD ':'RD', 'RDO ':'RDO', 'RD1 ':'RD1', 'RD2 ':'RD2', 'VD ':'VD', 'FE ':'FE', 'LD ': 'LD', 'ARVD ': 'ARVD',
                           'ARVDO ': 'ARVDO', 'TI ': 'TI', 'TO ':'TI', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD', 'PT ': 'PT', 'TPT ':'TPT', 'CX ': 'CX'})
        
                        df = st.session_state.df[['ART','AS', 'AG','VD', 'RD','GD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'RDO', 'ARVD', 'ARVDO', 'TPT', 'CX','PT']].copy()
                        df['ART'] = df['ART'].astype(str)
                        df['A'] = df['ART'].str.replace('[^0-9]', '', regex=True)
                        df['A'] = pd.to_numeric(df['A'], errors= 'coerce')
                        df = df[df['A']>0].copy()
                        #df.dropna(subset='ART', inplace=True)
                        
                        df[['AS', 'RD', 'VD','TO','TI']] = df[['AS', 'RD', 'VD','TO','TI']].astype(str)
                        if df['TI'].str.contains('YES').any():
                            st.write("You may be using the Transfer in column instead of the Transfer_in Obs date column")
                            st.stop()
                        
                        df['AS'] = df['AS'].astype(str)
                        df['ARVD'] = df['ARVD'].astype(str)
                        df['ARVDO'] = df['ARVDO'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['RDO'] = df['RDO'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)
                        
                        y = pd.DataFrame({'ART' :['2','3','4','5'], 'TI':['1-1-1',1,'1/1/1','3 8 2001'], 'RD':['1-1-1',1,'1/1/1','3 8 2001'],'DD':['1-1-1',1,'1/1/1','3 8 2001'], 
                                        'TO':['1-1-1',1,'1/1/1','3 8 2001'], 'AS':['1-1-1',1,'1/1/1','3 8 2001'], 'VD':['1-1-1',1,'1/1/1','3 8 2001'],'RD1':['1-1-1',1,'1/1/1','3 8 2001'],
                                        'RD2':['1-1-1',1,'1/1/1','3 8 2001'],'RDO':['1-1-1',1,'1/1/1','3 8 2001'], 'ARVD':['1-1-1',1,'1/1/1','3 8 2001'], 'ARVDO':['1-1-1',1,'1/1/1','3 8 2001'],
                                        'LD':['1-1-1',1,'1/1/1','3 8 2001'],'FE':['1-1-1',1,'1/1/1','3 8 2001']})   
        
                        
                        df['AS'] = df['AS'].astype(str)
                        df['ARVDO'] = df['ARVDO'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['RDO'] = df['RDO'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)
            
                        df['AS'] = df['AS'].str.replace('00:00:00', '', regex=True)
                        df['ARVDO'] = df['ARVDO'].str.replace('00:00:00', '', regex=True)
                        df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
                        df['RD1'] = df['RD1'].str.replace('00:00:00', '', regex=True)
                        df['RD2'] = df['RD2'].str.replace('00:00:00', '', regex=True)
                        df['RDO'] = df['RDO'].str.replace('00:00:00', '', regex=True)
                        df['TI'] = df['TI'].str.replace('00:00:00', '', regex=True)
                        df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
                        df['VD'] = df['VD'].str.replace('00:00:00', '', regex=True)
                        df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
                        df['LD'] = df['LD'].str.replace('00:00:00', '', regex=True)
                        df['FE'] = df['FE'].str.replace('00:00:00', '', regex=True)
            
                        df = pd.concat([df,y])
            
                        df['AS'] = df['AS'].astype(str) ###
                        df['ARVDO'] = df['ARVDO'].astype(str)
                        df['RD'] = df['RD'].astype(str) ###
                        df['RD1'] = df['RD1'].astype(str)##
                        df['RD2'] = df['RD2'].astype(str)##
                        df['RDO'] = df['RDO'].astype(str)
                        df['TI'] = df['TI'].astype(str) ##
                        df['TO'] = df['TO'].astype(str) ##
                        df['VD'] = df['VD'].astype(str) ###
                        df['DD'] = df['DD'].astype(str) ####
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)
            
            
                        # SPLITTING ART START DATE
                        A = df[df['AS'].str.contains('-')].copy()
                        a = df[~df['AS'].str.contains('-')].copy()
                        B = a[a['AS'].str.contains('/')].copy()
                        C = a[~a['AS'].str.contains('/')].copy()
                        E = C[C['AS'].str.contains(' ')].copy()
                        D = C[~C['AS'].str.contains(' ')].copy()
            
                        A[['Ayear', 'Amonth', 'Aday']] = A['AS'].str.split('-', expand = True)
                        B[['Ayear', 'Amonth', 'Aday']] = B['AS'].str.split('/', expand = True)
                        try:
                            D['AS'] = pd.to_numeric(D['AS'], errors='coerce')
                            D['AS'] = pd.to_datetime(D['AS'], origin='1899-12-30', unit='D', errors='coerce')
                            D['AS'] =  D['AS'].astype(str)
                            D[['Ayear', 'Amonth', 'Aday']] = D['AS'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['AS'] = pd.to_datetime(E['AS'],format='%d %m %Y', errors='coerce')
                            E['AS'] =  E['AS'].astype(str)
                            E[['Ayear', 'Amonth', 'Aday']] = E['AS'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E]) 
            
                        # SPLITTING DEATH DATE
                        A = df[df['DD'].str.contains('-')].copy()
                        a = df[~df['DD'].str.contains('-')].copy()
                        B = a[a['DD'].str.contains('/')].copy()
                        C = a[~a['DD'].str.contains('/')].copy()
                        E = C[C['DD'].str.contains(' ')].copy()
                        D = C[~C['DD'].str.contains(' ')].copy()
                        A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
                        B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)

                        try:
                            D['DD'] = pd.to_numeric(D['DD'], errors='coerce')
                            D['DD'] = pd.to_datetime(D['DD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['DD'] =  D['DD'].astype(str)
                            D[['Dyear', 'Dmonth', 'Dday']] = D['DD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['DD'] = pd.to_datetime(E['DD'],format='%d %m %Y', errors='coerce')
                            E['DD'] =  E['DD'].astype(str)
                            E[['Dyear', 'Dmonth', 'Dday']] = E['DD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E]) 
                    
                        # SORTING THE RETURN VISIT DATE
                        A = df[df['RD'].str.contains('-')].copy()
                        a = df[~df['RD'].str.contains('-')].copy()
                        B = a[a['RD'].str.contains('/')].copy()
                        C = a[~a['RD'].str.contains('/')].copy()
                        E = C[C['RD'].str.contains(' ')].copy()
                        D = C[~C['RD'].str.contains(' ')].copy()
                       
                        #D = C[C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
                        #E = C[~C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
                
                        A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
                        B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
                        try:
                            D['RD'] = pd.to_numeric(D['RD'], errors='coerce')
                            D['RD'] = pd.to_datetime(D['RD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD'] =  D['RD'].astype(str)
                            D[['Ryear', 'Rmonth', 'Rday']] = D['RD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD'] = pd.to_datetime(E['RD'],format='%d %m %Y', errors='coerce')
                            E['RD'] =  E['RD'].astype(str)
                            E[['Ryear', 'Rmonth', 'Rday']] = E['RD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E]) 
                    
                        #SORTING THE VD DATE
                        A = df[df['VD'].str.contains('-')].copy()
                        a = df[~df['VD'].str.contains('-')].copy()
                        B = a[a['VD'].str.contains('/')].copy()
                        C = a[~a['VD'].str.contains('/')].copy()
                        E = C[C['VD'].str.contains(' ')].copy()
                        D = C[~C['VD'].str.contains(' ')].copy()
            
                        A[['Vyear', 'Vmonth', 'Vday']] = A['VD'].str.split('-', expand = True)
                        B[['Vyear', 'Vmonth', 'Vday']] = B['VD'].str.split('/', expand = True)
                        try:
                            D['VD'] = pd.to_numeric(D['VD'], errors='coerce')
                            D['VD'] = pd.to_datetime(D['VD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['VD'] =  D['VD'].astype(str)
                            D[['Vyear', 'Vmonth', 'Vday']] = D['VD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['VD'] = pd.to_datetime(E['VD'],format='%d %m %Y', errors='coerce')
                            E['VD'] =  E['VD'].astype(str)
                            E[['Vyear', 'Vmonth', 'Vday']] = E['VD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                       
                        #SORTING THE TO DATE
                        A = df[df['TO'].str.contains('-')].copy()
                        a = df[~df['TO'].str.contains('-')].copy()
                        B = a[a['TO'].str.contains('/')].copy()
                        C = a[~a['TO'].str.contains('/')].copy()
                        E = C[C['TO'].str.contains(' ')].copy()
                        D = C[~C['TO'].str.contains(' ')].copy()
            
                        A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
                        B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
                        try:
                            D['TO'] = pd.to_numeric(D['TO'], errors='coerce')
                            D['TO'] = pd.to_datetime(D['TO'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TO'] =  D['TO'].astype(str)
                            D[['Tyear', 'Tmonth', 'Tday']] = D['TO'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TO'] = pd.to_datetime(E['TO'],format='%d %m %Y', errors='coerce')
                            E['TO'] =  E['TO'].astype(str)
                            E[['Tyear', 'Tmonth', 'Tday']] = E['TO'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
            
                    #SORTING THE TI DATE
                        A = df[df['TI'].str.contains('-')].copy()
                        a = df[~df['TI'].str.contains('-')].copy()
                        B = a[a['TI'].str.contains('/')].copy()
                        C = a[~a['TI'].str.contains('/')].copy()
                        E = C[C['TI'].str.contains(' ')].copy()
                        D = C[~C['TI'].str.contains(' ')].copy()
            
                        A[['Tiyear', 'Timonth', 'Tiday']] = A['TI'].str.split('-', expand = True)
                        B[['Tiyear', 'Timonth', 'Tiday']] = B['TI'].str.split('/', expand = True)
                        try:
                            D['TI'] = pd.to_numeric(D['TI'], errors='coerce')
                            D['TI'] = pd.to_datetime(D['TI'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TI'] =  D['TI'].astype(str)
                            D[['Tiyear', 'Timonth', 'Tiday']] = D['TI'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TI'] = pd.to_datetime(E['TI'],format='%d %m %Y', errors='coerce')
                            E['TI'] =  E['TI'].astype(str)
                            E[['Tiyear', 'Timonth', 'Tiday']] = E['TI'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
            
                        # SORTING THE RETURN VISIT DATE1
                        A = df[df['RD1'].str.contains('-')].copy()
                        a = df[~df['RD1'].str.contains('-')].copy()
                        B = a[a['RD1'].str.contains('/')].copy()
                        C = a[~a['RD1'].str.contains('/')].copy()
                        E = C[C['RD1'].str.contains(' ')].copy()
                        D = C[~C['RD1'].str.contains(' ')].copy()
                
                        A[['R1year', 'R1month', 'R1day']] = A['RD1'].str.split('-', expand = True)
                        B[['R1year', 'R1month', 'R1day']] = B['RD1'].str.split('/', expand = True)
                        try:
                            D['RD1'] = pd.to_numeric(D['RD1'], errors='coerce')
                            D['RD1'] = pd.to_datetime(D['RD1'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD1'] =  D['RD1'].astype(str)
                            D[['R1year', 'R1month', 'R1day']] = D['RD1'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD1'] = pd.to_datetime(E['RD1'],format='%d %m %Y', errors='coerce')
                            E['RD1'] =  E['RD1'].astype(str)
                            E[['R1year', 'R1month', 'R1day']] = E['RD1'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                    
                        # SORTING THE RETURN VISIT DATE2
                        A = df[df['RD2'].str.contains('-')].copy()
                        a = df[~df['RD2'].str.contains('-')].copy()
                        B = a[a['RD2'].str.contains('/')].copy()
                        C = a[~a['RD2'].str.contains('/')].copy()
                        E = C[C['RD2'].str.contains(' ')].copy()
                        D = C[~C['RD2'].str.contains(' ')].copy()
                
                        A[['R2year', 'R2month', 'R2day']] = A['RD2'].str.split('-', expand = True)
                        B[['R2year', 'R2month', 'R2day']] = B['RD2'].str.split('/', expand = True)
                        try:
                            D['RD2'] = pd.to_numeric(D['RD2'], errors='coerce')
                            D['RD2'] = pd.to_datetime(D['RD2'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD2'] =  D['RD2'].astype(str)
                            D[['R2year', 'R2month', 'R2day']] = D['RD2'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD2'] = pd.to_datetime(E['RD2'],format='%d %m %Y', errors='coerce')
                            E['RD2'] =  E['RD2'].astype(str)
                            E[['R2year', 'R2month', 'R2day']] = E['RD2'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                    
                        # SORTING THE RETURN VISIT OBS DATE
                        A = df[df['RDO'].str.contains('-')].copy()
                        a = df[~df['RDO'].str.contains('-')].copy()
                        B = a[a['RDO'].str.contains('/')].copy()
                        C = a[~a['RDO'].str.contains('/')].copy()
                        E = C[C['RDO'].str.contains(' ')].copy()
                        D = C[~C['RDO'].str.contains(' ')].copy()
                
                        A[['ROyear', 'ROmonth', 'ROday']] = A['RDO'].str.split('-', expand = True)
                        B[['ROyear', 'ROmonth', 'ROday']] = B['RDO'].str.split('/', expand = True)
                        try:
                            D['RDO'] = pd.to_numeric(D['RDO'], errors='coerce')
                            D['RDO'] = pd.to_datetime(D['RDO'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RDO'] =  D['RDO'].astype(str)
                            D[['ROyear', 'ROmonth', 'ROday']] = D['RDO'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RDO'] = pd.to_datetime(E['RDO'],format='%d %m %Y', errors='coerce')
                            E['RDO'] =  E['RDO'].astype(str)
                            E[['ROyear', 'ROmonth', 'ROday']] = E['RDO'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
            
                        # SORTING THE LAST ENCOUNTER DATES
                        A = df[df['LD'].str.contains('-')].copy()
                        a = df[~df['LD'].str.contains('-')].copy()
                        B = a[a['LD'].str.contains('/')].copy()
                        C = a[~a['LD'].str.contains('/')].copy()
                        E = C[C['LD'].str.contains(' ')].copy()
                        D = C[~C['LD'].str.contains(' ')].copy()
                
                        A[['Lyear', 'Lmonth', 'Lday']] = A['LD'].str.split('-', expand = True)
                        B[['Lyear', 'Lmonth', 'Lday']] = B['LD'].str.split('/', expand = True)
                        try:
                            D['LD'] = pd.to_numeric(D['LD'], errors='coerce')
                            D['LD'] = pd.to_datetime(D['LD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['LD'] =  D['LD'].astype(str)
                            D[['Lyear', 'Lmonth', 'Lday']] = D['LD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['LD'] = pd.to_datetime(E['LD'],format='%d %m %Y', errors='coerce')
                            E['LD'] =  E['LD'].astype(str)
                            E[['Lyear', 'Lmonth', 'Lday']] = E['LD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                    
                        # SORTING THE ARV DISPENSED DATES
                        A = df[df['ARVDO'].str.contains('-')].copy()
                        a = df[~df['ARVDO'].str.contains('-')].copy()
                        B = a[a['ARVDO'].str.contains('/')].copy()
                        C = a[~a['ARVDO'].str.contains('/')].copy()
                        E = C[C['ARVDO'].str.contains(' ')].copy()
                        D = C[~C['ARVDO'].str.contains(' ')].copy()
                
                        A[['Aryear', 'Armonth', 'Arday']] = A['ARVDO'].str.split('-', expand = True)
                        B[['Aryear', 'Armonth', 'Arday']] = B['ARVDO'].str.split('/', expand = True)
                        try:
                            D['ARVDO'] = pd.to_numeric(D['ARVDO'], errors='coerce')
                            D['ARVDO'] = pd.to_datetime(D['ARVDO'], origin='1899-12-30', unit='D', errors='coerce')
                            D['ARVDO'] =  D['ARVDO'].astype(str)
                            D[['Aryear', 'Armonth', 'Arday']] = D['ARVDO'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['ARVDO'] = pd.to_datetime(E['ARVDO'],format='%d %m %Y', errors='coerce')
                            E['ARVDO'] =  E['ARVDO'].astype(str)
                            E[['Aryear', 'Armonth', 'Arday']] = E['ARVDO'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
            
                        # SORTING THE FIRST ENCOUNTER
                        A = df[df['FE'].str.contains('-')].copy()
                        a = df[~df['FE'].str.contains('-')].copy()
                        B = a[a['FE'].str.contains('/')].copy()
                        C = a[~a['FE'].str.contains('/')].copy()
                        E = C[C['FE'].str.contains(' ')].copy()
                        D = C[~C['FE'].str.contains(' ')].copy()
                
                        A[['Fyear', 'Fmonth', 'Fday']] = A['FE'].str.split('-', expand = True)
                        B[['Fyear', 'Fmonth', 'Fday']] = B['FE'].str.split('/', expand = True)
                        try:
                            D['FE'] = pd.to_numeric(D['FE'], errors='coerce')
                            D['FE'] = pd.to_datetime(D['FE'], origin='1899-12-30', unit='D', errors='coerce')
                            D['FE'] =  D['FE'].astype(str)
                            D[['Fyear', 'Fmonth', 'Fday']] = D['FE'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['FE'] = pd.to_datetime(E['FE'],format='%d %m %Y', errors='coerce')
                            E['FE'] =  E['FE'].astype(str)
                            E[['Fyear', 'Fmonth', 'Fday']] = E['FE'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
            
                        #BRINGING BACK THE / IN DATES
                        df['AS'] = df['AS'].astype(str)
                        df['ARVDO'] = df['ARVDO'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['RDO'] = df['RDO'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)
            
            #             #Clearing NaT from te dates
                        df['AS'] = df['AS'].str.replace('NaT', '',regex=True)
                        df['ARVDO'] = df['ARVDO'].str.replace('NaT', '',regex=True)
                        df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
                        df['RD1'] = df['RD1'].str.replace('NaT', '',regex=True)
                        df['RD2'] = df['RD2'].str.replace('NaT', '',regex=True)
                        df['RDO'] = df['RDO'].str.replace('NaT', '',regex=True)
                        df['TI'] = df['TI'].str.replace('NaT', '',regex=True)
                        df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
                        df['VD'] = df['VD'].str.replace('NaT', '',regex=True)
                        df['DD'] = df['DD'].str.replace('NaT', '',regex=True)
                        df['LD'] = df['LD'].str.replace('NaT', '',regex=True)
                        df['FE'] = df['FE'].str.replace('NaT', '',regex=True)
            
                                    #SORTING THE VIRAL LOAD YEARS
                    
                        df[['Vyear', 'Vmonth', 'Vday']] =df[['Vyear', 'Vmonth', 'Vday']].apply(pd.to_numeric, errors = 'coerce') 
                        df['Vyear'] = df['Vyear'].fillna(994)
                        a = df[df['Vyear']>31].copy()
                        b = df[df['Vyear']<32].copy()
                        b = b.rename(columns={'Vyear': 'Vday2', 'Vday': 'Vyear'})
                        b = b.rename(columns={'Vday2': 'Vday'})
                        df = pd.concat([a,b])
                        dfa = df.shape[0]
            
            
                        #SORTING THE TI YEARS
                        df[['Tiyear', 'Tiday']] =df[['Tiyear','Tiday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Tiyear'] = df['Tiyear'].fillna(994)
                        a = df[df['Tiyear']>31].copy()
                        b = df[df['Tiyear']<32].copy()
                        b = b.rename(columns={'Tiyear': 'Tiday2', 'Tiday': 'Tiyear'})
                        b = b.rename(columns={'Tiday2': 'Tiday'})
                        df = pd.concat([a,b])
                        dfb = df.shape[0]
            
                        # #SORTING THE RETURN VISIT DATE YEARS
                        df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        
                        df['Ryear'] = df['Ryear'].fillna(994)
                        a = df[df['Ryear']>31].copy()
                        b = df[df['Ryear']<32].copy()
                        b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
                        b = b.rename(columns={'Rday2': 'Rday'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
                        
                            #SORTING THE TRANSFER OUT DATE YEAR
                        df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
                        df['Tyear'] = df['Tyear'].fillna(994)
                        a = df[df['Tyear']>31].copy()
                        b = df[df['Tyear']<32].copy()
                        b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
                        b = b.rename(columns={'Tday2': 'Tday'})
                        df = pd.concat([a,b])         
                        
                        #SORTING THE ART START YEARS
                        df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Ayear'] = df['Ayear'].fillna(994)
                        a = df[df['Ayear']>31].copy()
                        b = df[df['Ayear']<32].copy()
                        b = b.rename(columns={'Ayear': 'Aday2', 'Aday': 'Ayear'})
                        b = b.rename(columns={'Aday2': 'Aday'})
                        df = pd.concat([a,b])
                        dfe = df.shape[0]
            
                        #SORTING THE ART START YEARS
                        df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Dyear'] = df['Dyear'].fillna(994)
                        a = df[df['Dyear']>31].copy()
                        b = df[df['Dyear']<32].copy()
                        b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
                        b = b.rename(columns={'Dday2': 'Dday'})
                        df = pd.concat([a,b])
                        dfe = df.shape[0]
            
                        # #SORTING THE RETURN VISIT DATE1
                        df[['R1day', 'R1year']] = df[['R1day', 'R1year']].apply(pd.to_numeric, errors='coerce')
                        
                        df['R1year'] = df['R1year'].fillna(994)
                        a = df[df['R1year']>31].copy()
                        b = df[df['R1year']<32].copy()
                        b = b.rename(columns={'R1year': 'R1day2', 'R1day': 'R1year'})
                        b = b.rename(columns={'R1day2': 'R1day'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
            
                        # #SORTING THE RETURN VISIT DATE2
                        df[['R2day', 'R2year']] = df[['R2day', 'R2year']].apply(pd.to_numeric, errors='coerce')
                        
                        df['R2year'] = df['R2year'].fillna(994)
                        a = df[df['R2year']>31].copy()
                        b = df[df['R2year']<32].copy()
                        b = b.rename(columns={'R2year': 'R2day2', 'R2day': 'R2year'})
                        b = b.rename(columns={'R2day2': 'R2day'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
            
                        # #SORTING THE RETURN VISIT OBS DATE
                        df[['ROday', 'ROyear']] = df[['ROday', 'ROyear']].apply(pd.to_numeric, errors='coerce')
                        
                        df['ROyear'] = df['ROyear'].fillna(994)
                        a = df[df['ROyear']>31].copy()
                        b = df[df['ROyear']<32].copy()
                        b = b.rename(columns={'ROyear': 'ROday2', 'ROday': 'ROyear'})
                        b = b.rename(columns={'ROday2': 'ROday'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
            
                        # #SORTING THE LAST ENCOUNTER
                        df[['Lday', 'Lyear']] = df[['Lday', 'Lyear']].apply(pd.to_numeric, errors='coerce')
                        
                        df['Lyear'] = df['Lyear'].fillna(994)
                        a = df[df['Lyear']>31].copy()
                        b = df[df['Lyear']<32].copy()
                        b = b.rename(columns={'Lyear': 'Lday2', 'Lday': 'Lyear'})
                        b = b.rename(columns={'Lday2': 'Lday'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
            
                        # #SORTING THE FIRST ENCOUNTER
                        df[['Fday', 'Fyear']] = df[['Fday', 'Fyear']].apply(pd.to_numeric, errors='coerce')
                        
                        df['Fyear'] = df['Fyear'].fillna(994)
                        a = df[df['Fyear']>31].copy()
                        b = df[df['Fyear']<32].copy()
                        b = b.rename(columns={'Fyear': 'Fday2', 'Fday': 'Fyear'})
                        b = b.rename(columns={'Fday2': 'Fday'})
            
                        df = pd.concat([a,b])
                        dfc = df.shape[0]
            
                        # #SORTING THE FIRST ENCOUNTER
                        df[['Arday', 'Aryear']] = df[['Arday', 'Aryear']].apply(pd.to_numeric, errors='coerce')
                        
                        df['Aryear'] = df['Aryear'].fillna(994)
                        a = df[df['Aryear']>31].copy()
                        b = df[df['Aryear']<32].copy()
                        b = b.rename(columns={'Aryear': 'Arday2', 'Arday': 'Aryear'})
                        b = b.rename(columns={'Arday2': 'Arday'})
                        df = pd.concat([a,b])
                        dfc = df.shape [0]
            
                        #CREATE WEEKS 
                        df['Rdaya'] = df['Rday'].astype(str).str.split('.').str[0]
                        df['Rmontha'] = df['Rmonth'].astype(str).str.split('.').str[0]
                        df['Ryeara'] = df['Ryear'].astype(str).str.split('.').str[0]
            
                        df['RETURN DATE'] = df['Rdaya'] + '/' + df['Rmontha'] + '/' + df['Ryeara']
                        df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
                        #CREATING WEEEK FOR RETURN VISIT DATE
                        df['RWEEK'] = df['RETURN DATE'].dt.strftime('%V')
                        df['RWEEK'] = pd.to_numeric(df['RWEEK'], errors='coerce')
                        df['RWEEK1'] = df['RWEEK'] + 13
                        #       #PARAMETERS FOR CIRA
                           
                        df['R1aya'] = df['R1day'].astype(str).str.split('.').str[0]
                        df['R1montha'] = df['R1month'].astype(str).str.split('.').str[0]
                        df['R1yeara'] = df['R1year'].astype(str).str.split('.').str[0]
                        df['RETURN DATE1'] = df['R1aya'] + '/' + df['R1montha'] + '/' + df['R1yeara']
                        df['RETURN DATE1'] = pd.to_datetime(df['RETURN DATE1'], format='%d/%m/%Y', errors='coerce')

                       #LAST ENCOUTER TO DATES
                        df['Ldaya'] = df['Lday'].astype(str).str.split('.').str[0]
                        df['Lmontha'] = df['Lmonth'].astype(str).str.split('.').str[0]
                        df['Lyeara'] = df['Lyear'].astype(str).str.split('.').str[0]
                        df['LAST DATE'] = df['Ldaya'] + '/' + df['Lmontha'] + '/' + df['Lyeara']
                        df['LAST DATE'] = pd.to_datetime(df['LAST DATE'], format='%d/%m/%Y', errors='coerce')
   
                        # df['RWEEKR'] = df['RETURN DATE1'].dt.strftime('%V') #Use R since 1 was already used
                        # df['RWEEKR'] = pd.to_numeric(df['RWEEKR'], errors='coerce')
                        # df['RWEEKR1'] = df['RWEEKR'] +13 #NOT NEEDED THIS Q SINCE WE ARE USING 
        
                        df['DURA'] = round((df['LAST DATE']-df['RETURN DATE1']).dt.days/30)
        
                        today = dt.date.today() 
                        todayr = pd.to_datetime(today)
                        #df['DURL'] = round((todayr)-(df['RETURN DATE']))#.dt.days / 30)  
                        df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'])
        
                        # Calculate the difference in months
                        df['DURL'] = round((todayr - df['RETURN DATE']).dt.days/30)
                        df['TODAY'] = todayr
                        dfaa = df[['A', 'TODAY','RETURN DATE', 'DURL']].copy()
            
                        def cira(a):
                            if a<1:
                                return 'UK'
                            elif a< 3:
                                return '<3 MTHS'
                            elif a <6:
                                return '3-5 MTHS'
                            elif a >5:
                                return '6 MTHS+'
                            else:
                                return 'UK'
                        df[['DURA', 'DURL']] = df[['DURA', 'DURL']].apply(pd.to_numeric, errors='coerce')
                        df['CIRAA'] = df['DURA'].apply(cira)
                        df['CIRAL'] = df['DURL'].apply(cira)
                        def ager(a):
                            if a< 1:
                                return '<01'
                            elif a < 10:
                                return '01 to 09'
                            elif a < 20:
                                return '10 to 19'
                            elif a < 30:
                                return '20-29'
                            elif a < 40:
                                return '30-39'
                            elif a < 50:
                                return '40-49'
                            elif a >49:
                                return '50+'
                        df['AG'] = pd.to_numeric(df['AG'], errors = 'coerce')
                        df['BAND'] = df['AG'].apply(ager)
                        yyy = df.copy()
        
                        #COPY FOR ONE YEAR BEFORE GETTING POT CURR
                        oneyear = df.copy()
                        #tttt = df.copy()
                        nsps = df.copy()
                        yyyu = df.copy()
                        line = df.copy()
                        missed = df.copy()
                       
                        missy = df.copy()
                        #yyy = df.copy()
        
                        #df['GROUP'] = df['AG'].apply(ager)
                        # Q1'S TXML ALTER (LAST Q'S TXML
                        df['Tyear'] = pd.to_numeric(df['Tyear'],errors='coerce')
                        last4 = df[df['Tyear']==994].copy()
                        last4['Dyear'] = pd.to_numeric(last4['Dyear'],errors='coerce')
                        last4 = last4[last4['Dyear']==994].copy()
                        last4[['Ryear', 'Rmonth']] = last4[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                        last4 = last4[((last4['Ryear']==2024) & (last4['Rmonth'].isin([9,10,11,12])))].copy() #CHANGED
                        last4[['Rmonth', 'Rday']] = last4[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                        last4 = last4[((last4['Rmonth']<12) | ((last4['Rmonth']==12) & (last4['Rday']<4)))].copy()
                        ciraa = last4.copy()
                        lastq4 = last4.shape[0]

                        # Q4'S TXML ALTER (PREVIOUS 2 Q'S TXML)
                        df['Tyear'] = pd.to_numeric(df['Tyear'],errors='coerce')
                        last3 = df[df['Tyear']==994].copy()
                        last3['Dyear'] = pd.to_numeric(last3['Dyear'],errors='coerce')
                        last3 = last3[last3['Dyear']==994].copy()
                        last3[['Ryear', 'Rmonth']] = last3[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                        last3 = last3[((last3['Ryear']==2024) & (last3['Rmonth'].isin([6,7,8])))].copy()
                        last3[['Rmonth', 'Rday']] = last3[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                        last3 = last3[((last3['Rmonth']>6) | ((last3['Rmonth']==6) & (last3['Rday']>2)))].copy()
                        cirab = last3.copy()
                        lastq3 = last3.shape[0]

                        #st.write(ciraa)
                        #st.write(cirab)

                        cira1 = pd.concat([ciraa, cirab], axis=0)
        
                        #POTENTIAL TXCUR ALTER... 
                        df[['Rmonth', 'Rday', 'Ryear']] = df[['Rmonth', 'Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        df25 = df[df['Ryear']>2024].copy()
                        df24 = df[df['Ryear'] == 2024].copy()
                        df24[['Rmonth', 'Rday']] = df24[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                        df24 = df24[((df24['Rmonth']>12) | ((df24['Rmonth']==12) & (df24['Rday']>3)))].copy()
                        df = pd.concat([df25, df24]).copy()
            
                        #REMOVE TO of the last reporting month
                        df[ 'Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
                        dfto = df[df['Tyear']!=994].copy()
                        dfnot = df[df['Tyear'] == 994].copy()
                        dfto[['Ryear', 'Rmonth']] = dfto[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                        dfto = dfto[((dfto['Ryear']!=2024) |((dfto['Ryear']==2024) & (dfto['Rmonth']!=12)))].copy() #OTHERS WOULD BE FALSE TOs, even those made last Q since they were brought as false if their RRDs were this year
                        df = pd.concat([dfto,dfnot])
        
                        #REMOVE the dead of the reporting month
                        df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
                        dfdd = df[df['Dyear']!=994].copy()
                        dfnot = df[df['Dyear'] == 994].copy()
                        #THOSE WHO DIED BEFORE FIRST MONTH OF THE Q
                        dfdd[['Dyear', 'Dmonth']] = dfdd[['Dyear', 'Dmonth']].apply(pd.to_numeric, errors='coerce')
                        dfdd = dfdd[((dfdd['Dyear']>2024) |((dfdd['Dyear']==2024) & (dfdd['Dmonth']>12)))].copy() #DOESN'T MAKE SENSE
                        df = pd.concat([dfdd,dfnot])
                        # Filter out empty DataFrames
                        #dfs = [dfi for dfi in [dfdd, dfnot] if not dfi.empty]
                        #df = pd.concat(dfs)
                        pot = df.shape[0] #THIS IS THE POTENTIAL TXCURR
                        #yyy = df.copy()

                        #QUARTERLY TX ML
                        dfcurr = df.copy()
                        #DEAD
                        dfcurr['Dyear'] = pd.to_numeric(dfcurr['Dyear'], errors='coerce')
                        deadq = dfcurr[dfcurr['Dyear']!=994].copy()  #THE DEAD
                        dfcurr = dfcurr[dfcurr['Dyear']==994].copy() #REMOVED THE DEAD
                        
                        ####TO
                        dfcurr['Tyear'] = pd.to_numeric(dfcurr['Tyear'], errors='coerce')
                        dfcurra = dfcurr[dfcurr['Tyear']==994].copy()  #NO TO 
                        
                        dfcto = dfcurr[dfcurr['Tyear']!=994].copy() #HAS TOs and no TOs
                        
                        dfcto['Ryear'] = pd.to_numeric(dfcto['Ryear'], errors = 'coerce')
                        dfcto[['Ryear', 'Rmonth']] =  dfcto[['Ryear', 'Rmonth']].apply(pd.to_numeric)
                        dfctoF = dfcto[ ((dfcto['Ryear']> 2025) | ((dfcto['Ryear'] ==2025) & (dfcto['Rmonth']>3))) ].copy()
                        dfctoT = dfcto[ ((dfcto['Ryear']< 2025) | ((dfcto['Ryear'] ==2025) & (dfcto['Rmonth']<4))) ].copy()
                        #OLD TOs BUT active in reporting month
                        dfctoc = dfcto[((dfcto['Ryear']==2025) & (dfcto['Rmonth']==3))].copy()
                        dfctoc['Rday'] = pd.to_numeric(dfctoc['Rday'], errors='coerce')
                        dfctoc = dfctoc[dfctoc['Rday']>3].copy()

                        dfcur = pd.concat([dfcurra, dfctoF, dfctoc])
                    
                        lacks = dfcur[((dfcur['Vyear']< 2024) | ((dfcur['Vyear'] ==2024) & (dfcur['Vmonth']<4)))]
                        lacks[['Ayear', 'Amonth']] = lacks[['Ayear', 'Amonth']].apply(pd.to_numeric, errors ='coerce')
                        lacks = lacks[((lacks['Ayear']<2024) |((lacks['Ayear']==2024)& (lacks['Amonth'] <10)))].copy()
                        lacks = lacks[lacks['Ayear']!=994].copy()
        
                        dfcur[['Ryear', 'Rmonth', 'Rday']] = dfcur[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors ='coerce')
                        curlosta = dfcur[dfcur['Ryear']< 2025].copy() #LOST IN DEC, MAY NOT APPLY NEXT Q
                        
                        curlostb = dfcur[dfcur['Ryear'] == 2025].copy() #LOST THIS YEAR
                        
                        curlostb[['Ryear', 'Rmonth', 'Rday']] = curlostb[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors ='coerce')
                        curlostc = curlostb[ ((curlostb['Rmonth']<3) |(( curlostb['Rmonth']==3) & (curlostb['Rday']<4)))].copy()
                        currlost = pd.concat([curlosta, curlostc])
        
                        cur26 = dfcur[dfcur['Ryear'] >2025].copy() #ACTIVE NEXT OTHER YEARS
                        cur25 = dfcur[dfcur['Ryear'] == 2025].copy() # ACTIVE THIS YEAR
                        
                        cur25[['Ryear', 'Rmonth', 'Rday']] = curlostb[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors ='coerce')
                        cur25 = cur25[ ((cur25['Rmonth']>3) |(( cur25['Rmonth']==3) & (curlostb['Rday']>3)))].copy()
                        dfcur = pd.concat([cur25, cur26])

                        #VL SECTION 
                        dfcur[['Vyear', 'Vmonth']] = dfcur[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        
                        has = dfcur[((dfcur['Vyear'] ==2025) | ((dfcur['Vyear'] ==2024) & (dfcur['Vmonth']>3)))]
        
                        #ONE YEAR COHORT
                        oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                        new = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([1,2,3])))].copy()
                        newtotal = new.shape[0]
            
                        new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                        tin = new[new['Tiyear']!=994].copy()
                        #one =new.shape[0]
                        newti = tin.shape[0]
                        orig = int(newtotal)-int(newti)

                        #LOSSES
                        new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                        newdead = new[new['Dyear']!=994].copy()
            
                        deadnew = newdead.shape[0]
                        new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD

                        #ACTIVE 1 YEAR
                        new['A'] = pd.to_numeric(new['A'], errors='coerce')
                        dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                        newcur = new[new['A'].isin(dfcur['A'])].copy()
                        newack = newcur.shape[0]
        
                        newlost = new[~new['A'].isin(dfcur['A'])].copy()
                        newlost['Tyear'] = pd.to_numeric(newlost['Tyear'], errors='coerce')
                        newlosto = newlost[newlost['Tyear']!=994].copy()
                        newlostout = newlosto.shape[0]
                        
                        newlost = newlost[newlost['Tyear']==994].copy()
                        newltfu = newlost.shape[0]
                        


                                #SIX YEAR COHORT
                        oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                        six = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([7,8,9])))].copy()
                        sixtotal = six.shape[0]
            
                        six[['Tiyear']] = six[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                        sixtin = six[six['Tiyear']!=994].copy()
                        
                        sixti = sixtin.shape[0]
                        origsix = int(sixtotal)-int(sixti)

                        #LOSSES
                        six['Dyear'] = pd.to_numeric(six['Dyear'], errors='coerce')
                        sixdead = six[six['Dyear']!=994].copy()
            
                        deadsix = sixdead.shape[0]
                        six = six[six['Dyear']==994].copy() #AFTER REMOVING THE DEAD
        
                        six['Tyear'] = pd.to_numeric(six['Tyear'], errors='coerce')

                        #ACTIVE 6
                        six['A'] = pd.to_numeric(six['A'], errors='coerce')
                        dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                        sixcur = six[six['A'].isin(dfcur['A'])].copy()
                        sixack = sixcur.shape[0]
        
                        sixlost = six[~six['A'].isin(dfcur['A'])].copy()
                        sixlost['Tyear'] = pd.to_numeric(sixlost['Tyear'], errors='coerce')
                        sixlosto = sixlost[sixlost['Tyear']!=994].copy()
                        sixlostout = sixlosto.shape[0]
                        sixlost = sixlost[sixlost['Tyear']==994].copy()
                        sixltfu = sixlost.shape[0]
                        list8 = [orig, newti, newtotal, deadnew, newlostout, newltfu,newack, origsix, sixti, sixtotal, deadsix, sixlostout, sixltfu,sixack,]

        
                        #MEASURES 
                        txcurr = dfcur.shape[0]
                        curto = dfctoT.shape[0]
                        hasvl = has.shape[0]
                        deadcur = deadq.shape[0]                       
                        lostq = currlost.shape[0]
                        vlcov = round((hasvl/txcurr)*100)
                        #TRANSFER OUTS
                        
                        #TRANSFER INS
                        df[['Tiyear', 'Timonth']] = df[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
                        dfti = df[((df['Tiyear']==2025) & (df['Timonth']>0))].copy() #TI MSUT HAVE BEEN MADE THIS YEAR
                        ti = dfti.shape[0]
            
                        # dfnot = df[((df['Tiyear']!=2024) | ((df['Tiyear']==2024) & (df['Timonth']<10)))].copy() #NO TI
                        # noti = dfnot.shape[0]  I DO NEED THESE ANYWHERE
            
                        #TX NEW THIS Q
                        dfnot[['Ayear', 'Amonth']] = dfnot[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        dfnew = dfnot[((dfnot['Ayear']==2025) & (dfnot['Amonth']>0))].copy() #TX NEWS MUST HAVE HAPPENED THIS YEAR ALSO
                        txnew = dfnew.shape[0]
                        #pppp= dfnew.copy()
                        dfold = dfnot[((dfnot['Ayear']!=2025) | ((dfnot['Ayear']==2025) & (dfnot['Amonth']<0)))].copy() #NO TI
                        dfcheck = dfold.copy() #use this to determine unknown gain (SITEGEDDE)
                        old = dfold.shape[0]
        
                        ##RTT
                        #RTT BY LAST ENCOUNTER to include only months in the reporting Q
                        dfold['Lyear'] = pd.to_numeric(dfold['Lyear'], errors='coerce') 
                        dfRTT = dfold[dfold['Lyear']==2025].copy() #ALTER
                        dfRTT['Lmonth'] = pd.to_numeric(dfRTT['Lmonth'], errors='coerce') 
                        dfRTT = dfRTT[dfRTT['Lmonth'].isin([1,2,3])].copy() #ALTER
            
                        #BY FIRST ENCOUNTER, To remove those first encountered in the Q
                        dfRTT['Fyear'] = pd.to_numeric(dfRTT['Fyear'], errors='coerce') 
                        dfRTTa = dfRTT[dfRTT['Fyear']==2025].copy() #ALTER
                        dfRTTb = dfRTT[dfRTT['Fyear']!=2025].copy() #ALTER
                        #BY FIRST ENCOUNTER
                        dfRTTa['Fmonth'] = pd.to_numeric(dfRTTa['Fmonth'], errors='coerce') 
                        dfRTTa = dfRTTa[~dfRTTa['Fmonth'].isin([1,2,3])].copy() # ALTER
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
      
              
                        #BY RD OBS DATE,  remove those that fall in the previous reporting Quarter
                        dfRTT['ROyear'] = pd.to_numeric(dfRTT['ROyear'], errors='coerce')
                        dfRTTa = dfRTT[dfRTT['ROyear']>2024].copy()
                        dfRTTb = dfRTT[dfRTT['ROyear']==2024].copy() 
                        dfRTTb[['ROmonth', 'ROday']] = dfRTTb[['ROmonth', 'ROday']].apply(pd.to_numeric, errors='coerce')
                        dfRTTb = dfRTTb[((dfRTTb['ROmonth']>12) | ((dfRTTb['ROmonth']==12) & (dfRTTb['ROday']>3)))].copy()
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
            
                        #BY RDDATE1,  take those that fall in the previous reporting Quarter
                        dfRTT['R1year'] = pd.to_numeric(dfRTT['R1year'], errors='coerce') 
                        dfRTTa = dfRTT[dfRTT['R1year']<2024].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb = dfRTT[dfRTT['R1year']==2024].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb[['R1month', 'R1day']] = dfRTTb[['R1month', 'R1day']].apply(pd.to_numeric, errors='coerce')
                        dfRTTb = dfRTTb[((dfRTTb['R1month']<12) | ((dfRTTb['R1month']==12) & (dfRTTb['R1day']<4)))].copy()
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
            
                        #BY RD DATE2,  take those that fall in the previous reporting Quarter
                        dfRTT['R2year'] = pd.to_numeric(dfRTT['R2year'], errors='coerce')
                        dfRTTa = dfRTT[dfRTT['R2year']<2024].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb = dfRTT[dfRTT['R2year']==2024].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb[['R2month', 'R2day']] = dfRTTb[['R2month', 'R2day']].apply(pd.to_numeric, errors='coerce')
                        dfRTTb = dfRTTb[((dfRTTb['R2month']<12) | ((dfRTTb['R2month']==12) & (dfRTTb['R2day']<4)))].copy()
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
            
                        #BY ARV DISPENSED, to take those that got ART in the Q
                        dfRTT['Aryear'] = pd.to_numeric(dfRTT['Aryear'], errors='coerce') 
                        dfRTT = dfRTT[dfRTT['Aryear']==2024].copy() 
                        dfRTT['Armonth'] = pd.to_numeric(dfRTT['Armonth'], errors='coerce') 
                        dfRTT = dfRTT[dfRTT['Armonth'].isin([1,2,3])].copy()
 
                        
            #######LOSSES. START FROM POTENTIAL CURR
                    #TRANSFER OUTS
                        
                        df['Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
                        dfnot = df[df['Tyear']==994].copy()
                        dfto = df[df['Tyear']!=994].copy()
                        wk = int(wk)
            
                        #FALSE TO OUTS BASED ON CURRENT WEEK
                        dfto[['Ryear', 'RWEEK1']] =  dfto[['Ryear', 'RWEEK1']].apply(pd.to_numeric, errors='coerce')
                        dfw = dfto[((dfto['Ryear']>2025) | ((dfto['Ryear']==2025) & (dfto['RWEEK1']>=wk)))].copy() #FALSE
                        false = dfw.shape[0]
                        dft = dfto[((dfto['Ryear']<2025) | ((dfto['Ryear']==2025) & (dfto['RWEEK1']<wk)))].copy()  ##TRUE
                        
                        true = dft.shape[0]
                        #add the false back to txcur
                        df = pd.concat([dfnot,dfw]) #WILL USE THIS FOR ACTIVE LATER

                        #THOSE THAT HAVE DIED SO FAR
                        df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
                        died = df[df['Dyear']!=994].copy() #DIED
                        dead = died.shape[0]
            
                        #THIS CURR WILL HAVE NO DEAD AND TRUE TO
                        df = df[df['Dyear'] == 994].copy() #LIVING, NO DEATH DATE
            
                        #REMOVNG CURRENT LOST
                        #USE CALENDAR WEEK FOR THIS Q, SWITCH TO SURGE WEEK NEXT Q
                        #lost 2 weeks
                        wk = int(wk)
                        wk2 = wk-1
                        wk3 = wk-2
                        wk4 = wk-3
        
                        df[['Ryear','Rmonth']] = df[['Ryear','Rmonth']].apply(pd.to_numeric,errors='coerce')
                        #df24 = df[((df['Ryear'] ==2024) & (df['Rmonth']<12))].copy()
                        df24a = df[df['Ryear'] ==2024].copy() #ALL 2024 ARE LOST SO ADD THEM TO ANY WEEK
                        df25 = df[df['Ryear'] ==2025].copy()
        
                        #df25 = df[((df['Ryear']>2024)| ((df['Ryear'] ==2024) & (df['Rmonth']==12)))].copy()
                        df25['RWEEK1'] = pd.to_numeric(df25['RWEEK1'], errors='coerce')
        
                        dfactive25 =df25[df25['RWEEK1']>=wk2] #still active within 2 weeks, only those of 2025 are considered, to avoid weeks of 2024
                  
                        #LOST IN TWO WEEKS... REAL MISSED APPOINTMENT FOR THIS (ADD ON THOSE OF 2024)
                        df2wksa =df25[df25['RWEEK1']<wk2].copy()
                        df2wks = pd.concat([df2wksa, df24a])  #WON'T BE NEEDED NEXT Qtr
            
                        
                        cira2 = df2wks.copy()
                        two = df2wks.shape[0]

                        df3wksa =df25[df25['RWEEK1']<wk3].copy()
                        df3wks = pd.concat([df3wksa, df24a])  #WON'T BE NEEDED NEXT Qtr
                        three = df3wks.shape[0]
                        #yyy3 = df3wks.copy()
            
                        df4wksa =df25[df25['RWEEK1']<wk4].copy()
                        df4wks = pd.concat([df4wksa, df24a])  #WON'T BE NEEDED NEXT Qtr
                        four = df4wks.shape[0]
                        #yyy4 = df4wks.copy()
            
                        #dfactive = pd.concat([dfactive24, df25]) #COMBINE THOSE ACTIVE IN TWO WEEKS AND THOSE OF 2025
                        dfactive = dfactive25.copy()
                        yyyuu  = dfactive.copy()
                        curr = dfactive.shape[0]
                        dfRTT['A'] = pd.to_numeric(dfRTT['A'], errors='coerce')
                        dfactive['A'] = pd.to_numeric(dfactive['A'], errors='coerce')
        
                        #RTT VS DFACTIVE
                        dfRTT = dfRTT[dfRTT['A'].isin(dfactive['A'])].copy()
                        rtt = dfRTT.shape[0]
                        #pppp=dfRTT.copy()
                        cactive = dfactive.copy()
                        clost = pd.concat([cira1,cira2])
        
        
                       #OF THOSE ACTIVE, HOW MANY WERE ON APPT 2 WEEKS AGO, 
                        dfactive['RWEEK1'] = pd.to_numeric(dfactive['RWEEK1'], errors='coerce')
                        appt = dfactive[dfactive['RWEEK1']<wk2].copy()
                        onappt = appt.shape[0]
    
                        #MMD AMONGST ACTIVE CLIENTS
                        dfactive['ARVD'] = dfactive['ARVD'].fillna(20)
                        dfactive['ARVD'] = pd.to_numeric(dfactive['ARVD'], errors='coerce')
                        def mmd(a):
                            if a<90:
                                return '<3 MTHS'
                            elif a< 180:
                                return '<6 MTHS'
                            else:
                                return '6 MTHS+'
                        dfactive = dfactive.copy() #avoid fragmentation
                        #pppp= dfactive.copy()
                        dfactive['MULTI'] = dfactive['ARVD'].apply(mmd)
                        dfactive['MULTI'] = dfactive['MULTI'].astype(str)
                        df2mths =  dfactive[dfactive['MULTI']=='<3 MTHS'].copy()
                        M2 = df2mths.shape[0]
                        df3mths =  dfactive[dfactive['MULTI']=='<6 MTHS'].copy()
                        M3 = df3mths.shape[0]
                        df6mths =  dfactive[dfactive['MULTI']=='6 MTHS+'].copy()
                        M6 = df6mths.shape[0]
            
                        #VL SECTION
                        #REMOVING SIX MONTHS TX NEW, to take those that got ART in the Q
                        dfactive['Ayear'] = pd.to_numeric(dfactive['Ayear'], errors='coerce') 
                        VLa = dfactive[dfactive['Ayear']<2024].copy()
                        VLb = dfactive[dfactive['Ayear']==2024].copy()
                        VLb = VLb[VLb['Amonth']<10].copy()
                        VL = pd.concat([VLa,VLb])
                        el = VL.shape[0]

        
                        VL[['Vyear', 'Vmonth']] = VL[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        WVL = VL[((VL['Vyear']>2024) | ((VL['Vyear']==2024) & (VL['Vmonth']>3)))].copy()
                        NVL = VL[((VL['Vyear']<2024) | ((VL['Vyear']==2024) & (VL['Vmonth']<4)))].copy()
                        nvl = NVL.shape[0]
                        wvl = WVL.shape[0]
            
                        #VL COV AMONG LOST CLIENTS
                        df2wks['Ayear'] = pd.to_numeric(df2wks['Ayear'], errors='coerce')
                        LVLa = df2wks[df2wks['Ayear']<2024].copy()
        
                        LVLb = df2wks[df2wks['Ayear']==2024].copy()
                        LVLb = LVLb[LVLb['Amonth']<10].copy()
                        LVL = pd.concat([LVLa,LVLb])
                        Lel = LVL.shape[0] #LOST ELIGIBLE
                        LVL[['Vyear', 'Vmonth']] = LVL[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        LWVL = LVL[((LVL['Vyear']>2024) | ((LVL['Vyear']==2024) & (LVL['Vmonth']>3)))].copy()
                        LNVL = LVL[((LVL['Vyear']<2024) | ((LVL['Vyear']==2024) & (LVL['Vmonth']<4)))].copy()
                        lnvl = LNVL.shape[0]
                        lwvl = LWVL.shape[0]
                        totalvl = pd.concat([LNVL,NVL])
    
                        #EARLY RETENTION
                        # #ONE YEAR COHORT
                        # oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                        # new = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([1,2,3])))].copy()
                        # newtotal = new.shape[0]
            
                        # new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                        # tin = new[new['Tiyear']!=994].copy()
                        # #one =new.shape[0]
                        # newti = tin.shape[0]
                        # orig = int(newtotal)-int(newti)

                        # #LOSSES
                        # new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                        # newdead = new[new['Dyear']!=994].copy()
            
                        # deadnew = newdead.shape[0]
                        # new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD
                        # new['Tyear'] = pd.to_numeric(new['Tyear'], errors='coerce')
                        
                        # newto = new[new['Tyear']!=994].copy()                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
                        # newto[['Ryear', 'RWEEK1']] = newto[['Ryear', 'RWEEK1']].apply(pd.to_numeric, errors='coerce')
                        # newto = newto[((newto['Ryear'] ==2024)|((newto['Ryear'] ==2025) & (newto['RWEEK1']>wk)))].copy()
                        # outnew = newto.shape[0]

                        # new['Tyear'] = pd.to_numeric(new['Tyear'], errors='coerce')
                        # newtfa = new[new['Tyear']!=994].copy()                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                             
                        # newtfa[['Ryear', 'RWEEK1']] = newtfa[['Ryear', 'RWEEK1']].apply(pd.to_numeric, errors='coerce')
                        # newtfa = newtfa[((newtfa['Ryear'] ==2025) & (newtfa['RWEEK1']>wk))].copy()
                        # new = new[new['Tyear']==994].copy() #withou TO
                        # new = pd.concat([newtfa, new])
                        
                        # netnew = new.shape[0]
            
                        # new['A'] = pd.to_numeric(new['A'], errors = 'coerce')
                        # dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
                        
                        # activen = new[new['A'].isin(dfactive['A'])].copy()
                        # lostn = new[~new['A'].isin(dfactive['A'])].copy()           
            
                        # newactive = activen.shape[0]
                        # newlost = lostn.shape[0]
                        #st.write(newlost)
        
                        #VL SECTION AT ONE YEAR
                    #     activen[['Vyear', 'Vmonth']] = activen[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                    #     WVLa = activen[ ((activen['Vyear']==2025) & (activen['Vmonth'].isin([1,2,3])))].copy()
                    #     NVLa = activen[((activen['Vyear']<2025) | ((activen['Vyear']==2025) & (activen['Vmonth']<1)))].copy()
                    #     nvla = NVLa.shape[0]
                    #     wvla = WVLa.shape[0]
                                
                    #     #ret = newtotal - newlost
                    #     if netnew == 0:
                    #         rete = 0
                    #     elif newactive == 0:
                    #         rete = 0
                    #     else:
                    #         rete = round((newactive/netnew)*100)
                    #         #rete = f"{rete} %"
                    #     #9 MONTH COHORT
            
                    #     oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    #     new9 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([4,5,6])))].copy()
                    #     newtotal9 = new9.shape[0]
            
                    #     new9[['Tiyear']] = new9[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                    #     tin9 = new9[new9['Tiyear']!=994].copy()
                    #     #one =new.shape[0]
                    #     newti9 = tin9.shape[0]
                    #     orig9 = int(newtotal9)-int(newti9)
                    #     new9['Dyear'] = pd.to_numeric(new9['Dyear'], errors='coerce')
                    #     newdead9 = new9[new9['Dyear']!=994].copy()
            
                    #     deadnew9 = newdead9.shape[0]
                    #     new9 = new9[new9['Dyear']==994].copy() #AFTER REMOVING THE DEAD
            
                    #     new9['Tyear'] = pd.to_numeric(new9['Tyear'], errors='coerce')
                        
                    #     newto9 = new9[new9['Tyear']!=994].copy()
                    #     outnew9 = newto9.shape[0]
                        
                    #     new9 = new9[new9['Tyear']==994].copy() #withou TO
                    #     netnew9 = new9.shape[0]
            
                    #     new9['A'] = pd.to_numeric(new9['A'], errors = 'coerce')
                    #     dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
                        
                    #     active9 = new9[new9['A'].isin(dfactive['A'])].copy()
                    #     lostn9 = new9[~new9['A'].isin(dfactive['A'])].copy()
                    
                    #     newactive9 = active9.shape[0]
                    #     newlost9 = lostn9.shape[0]
                    #     #ret = newtotal - newlost
                    #     if netnew9 == 0:
                    #         rete9 = 0
                    #     elif newactive9 == 0:
                    #         rete9 = 0
                    #     else:
                    #         rete9 = round((newactive9/netnew9)*100)
                    
                    
            
                    # #6 MONTH COHORT
                    #     oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    #     new6 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([7,8,9])))].copy()
                    #     newtotal6 = new6.shape[0]
            
                    #     new6[['Tiyear']] = new6[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                    #     tin6 = new6[new6['Tiyear']!=994].copy()
                    #     #one =new.shape[0]
                    #     newti6 = tin6.shape[0]
                    #     orig6 = int(newtotal6)-int(newti6)
                    #     new6['Dyear'] = pd.to_numeric(new6['Dyear'], errors='coerce')
                    #     newdead6 = new6[new6['Dyear']!=994].copy()
        
                    #     deadnew6 = newdead6.shape[0]
                    #     new6 = new6[new6['Dyear']==994].copy() #AFTER REMOVING THE DEAD
            
                    #     new6['Tyear'] = pd.to_numeric(new6['Tyear'], errors='coerce')
                        
                    #     newto6 = new6[new6['Tyear']!=994].copy()
                    #     outnew6 = newto6.shape[0]
                        
                    #     new6 = new6[new6['Tyear']==994].copy() #withou TO
                    #     netnew6 = new6.shape[0]
            
                    #     new6['A'] = pd.to_numeric(new6['A'], errors = 'coerce')
                    #     dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
                        
                    #     active6 = new6[new6['A'].isin(dfactive['A'])].copy()
                    #     lostn6 = new6[~new6['A'].isin(dfactive['A'])].copy()
                    
                    #     newactive6 = active6.shape[0]
                    #     newlost6 = lostn6.shape[0]
                    #     #st.write(newlost)
                    #     #VL SECTION AT 6 MONTHS
                    #     active6[['Vyear', 'Vmonth']] = active6[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                    #     WVLa6 = active6[active6['Vyear']==2024 ].copy()
                    #     NVLa6 = active6[active6['Vyear']!=2024].copy()
                    #     nvla6 = NVLa6.shape[0]
                    #     wvla6 = WVLa6.shape[0]
                    #     #ret = newtotal - newlost
                    #     if netnew6 == 0:
                    #         rete6 = 0
                    #     elif newactive6 == 0:
                    #         rete6 = 0
                    #     else:
                    #         rete6 = round((newactive6/netnew6)*100)
                    #         #rete6 = f"{rete6} %"
            
                    # #3 MONTH COHORT
                    #     oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    #     new3 = oneyear[((oneyear['Ayear']==2024) & (oneyear['Amonth'].isin([10,11,12])))].copy()
                    #     newtotal3 = new3.shape[0]
            
                    #     new3[['Tiyear']] = new3[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                    #     tin3 = new3[new3['Tiyear']!=994].copy()
                    #     #one =new.shape[0]
                    #     newti3 = tin3.shape[0]
                    #     orig3 = int(newtotal3)-int(newti3)
                        
                    #     new3['Dyear'] = pd.to_numeric(new3['Dyear'], errors='coerce')
                    #     newdead3 = new3[new3['Dyear']!=994].copy()
            
                    #     deadnew3 = newdead3.shape[0]
                    #     new3 = new3[new3['Dyear']==994].copy() #AFTER REMOVING THE DEAD
            
                    #     new3['Tyear'] = pd.to_numeric(new3['Tyear'], errors='coerce')
                        
                    #     newto3 = new3[new3['Tyear']!=994].copy()
                    #     outnew3 = newto3.shape[0]
                        
                    #     new3 = new3[new3['Tyear']==994].copy() #withou TO
                    #     netnew3 = new3.shape[0]
            
                    #     new3['A'] = pd.to_numeric(new3['A'], errors = 'coerce')
                    #     dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
                        
                    #     active3 = new3[new3['A'].isin(dfactive['A'])].copy()
                    #     lostn3 = new3[~new3['A'].isin(dfactive['A'])].copy()
                        
            
                    #     newactive3 = active3.shape[0]
                    #     newlost3 = lostn3.shape[0]
                    #     #st.write(newlost)
                                
                    #     #ret = newtotal - newlost
                    #     if netnew3 == 0:
                    #         rete3 = 0
                    #     elif newactive3 == 0:
                    #         rete3 = 0
                    #     else:
                    #         rete3 = round((newactive3/netnew3)*100)
                    #         #rete3 = f"{rete3} %"

                    #     #TX NEWS
                    #     oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    #     new1 = oneyear[((oneyear['Ayear']==2025) & (oneyear['Amonth'].isin([1,2,3])))].copy()
                    #     newtotal1 = new1.shape[0]
            
                    #     new1[['Tiyear']] = new1[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                    #     tin1 = new1[new1['Tiyear']!=994].copy()
                    #     #one =new.shape[0]
                    #     newti1 = tin1.shape[0]
                    #     orig1 = int(newtotal1)-int(newti1)
                        
                    #     new1['Dyear'] = pd.to_numeric(new1['Dyear'], errors='coerce')
                    #     newdead1 = new1[new1['Dyear']!=994].copy()
            
                    #     deadnew1 = newdead1.shape[0]
                    #     new1 = new1[new1['Dyear']==994].copy() #AFTER REMOVING THE DEAD
            
                    #     new1['Tyear'] = pd.to_numeric(new1['Tyear'], errors='coerce')
                        
                    #     newto1 = new1[new1['Tyear']!=994].copy()
                        
                    #     outnew1 = newto1.shape[0]
                        
                    #     new1 = new1[new1['Tyear']==994].copy() #withou TO
                    #     netnew1 = new1.shape[0]
            
                    #     new1['A'] = pd.to_numeric(new1['A'], errors = 'coerce')
                    #     dfactive['A'] = pd.to_numeric(dfactive['A'], errors = 'coerce')
                        
                    #     active1 = new1[new1['A'].isin(dfactive['A'])].copy()
                    #     lostn1 = new1[~new1['A'].isin(dfactive['A'])].copy() 

                    #     #FALSE TO IN TX NEWS YR
                    #     newto1['A'] = pd.to_numeric(newto1['A'], errors = 'coerce')
                    #     falseto = newto1[newto1['A'].isin(dfactive['A'])].copy()
                    #     trueto1 = newto1[~newto1['A'].isin(dfactive['A'])].copy() 

                    #     newto1 = trueto1.copy()
                    #     outnew1 = newto1.shape[0]
                    #     active1 = pd.concat([active1,falseto])
            
                    #     newactive1 = active1.shape[0]
                    #     newlost1 = lostn1.shape[0]
                    #     #st.write(newlost)
                    #     ret = newtotal - newlost
                    #     if netnew1 == 0:
                    #         rete1 = 0
                    #     elif newactive1 == 0:
                    #         rete1 = 0
                    #     else:
                    #         rete1 = round((newactive1/netnew1)*100)
                            #rete1 = f"{rete1} %"
                        # if st.session_state.reader:
                        #     st.write(pot)
                        # list1 = [lastq4,pot,ti,txnew,rtt,true,dead,two,three,four,curr,M2,M3,M6, onappt,lastq3] #TX
                        
                        # list2 = [curr,el,wvl,nvl,two,Lel, lnvl,lwvl, newactive,wvla,nvla,newactive6,wvla6,nvla6] #VL
                        
                        # list3 = [newtotal, orig,newti,deadnew,outnew, newlost,netnew,newactive,rete,
                        #              newtotal6, orig6,newti6,deadnew6,outnew6,newlost6,netnew6, newactive6,rete6,
                        #           newtotal9, orig9,newti9,deadnew9,outnew9,newlost9,netnew9, newactive9,rete9] #YEAR
                        # list4 = [newtotal3, orig3,newti3,deadnew3,outnew3,newlost3,netnew3, 
                        #              newactive3,rete3,newtotal1, orig1,newti1,deadnew1,outnew1,newlost1,netnew1, newactive1,rete1] #THRRE
                        # st.session_state.reader =True
                        lst = df2wks[['A', 'RD']].copy()
                        tout = dft[['A', 'TO']].copy()
                        die = died[['A', 'DD']].copy()
                        vir = totalvl[['A', 'VD']].copy()
                        #one = lostn[['A','AS','RD']].copy()
                    
                        lst['MISSED'] = np.nan
                        lst['MISSED'] = lst['MISSED'].fillna('MISSED APPT')
                        lst['A'] = pd.to_numeric(lst['A'], errors='coerce')
                    
                        tout['TRANSFERED'] = np.nan
                        tout['TRANSFERED'] = tout['TRANSFERED'].fillna('TO')
                        first = pd.concat([lst,tout])#, on = 'A', how = 'outer')
                    
                        die['DEAD?'] = np.nan
                        die['DEAD?'] = die['DEAD?'].fillna('DIED')
                        second = pd.concat([first,die])#, on = 'A', how = 'outer')
                        
                        vir['VL STATUS'] = np.nan
                        vir['VL STATUS'] = vir['VL STATUS'].fillna('DUE')
                        vir['A'] = pd.to_numeric(vir['A'], errors='coerce')
                        second['A'] = pd.to_numeric(second['A'], errors='coerce')
                        third = pd.merge(second,vir, on = 'A', how = 'outer')
                    
                        # one['ONE YEAR'] = np.nan
                        # one['ONE YEAR'] = one['ONE YEAR'].fillna('ONE YEAR IIT')
                        # forth = pd.concat([third,one])#, on = 'A', how = 'outer')
                            
    if st.session_state.reader:                                                    
        file2 = r'CLUSTERS.csv'
        dfx = pd.read_csv(file2)
        clusters  = list(dfx['CLUSTER'].unique())
        cluster = st.radio(label='**Choose your cluster**', options=clusters,index=None, horizontal=True)
        if not cluster:
            st.stop()
        else:
            districts = dfx[dfx['CLUSTER']==cluster]
            districts = list(districts['DISTRICT'].unique())
            district = st.radio(label='**Choose your district**', options=districts,index=None, horizontal=True)
            if not district:
                st.stop()
            else:
                facilities = dfx[dfx['DISTRICT']==district]
                facilities = facilities['FACILITY'].unique()
                facility = st.selectbox(label='**Choose your facility**', options=facilities,index=None)
                if not facility:
                    st.stop()
                else:
                    facy = facility
                    #st.session_state.fac = facility
                    if 'fac' not in st.session_state:
                           st.session_state.fac = facility
                    pass
                    if str(facy) != str(st.session_state.fac):
                            #st.info(f'DATA FOR {facy} NOT SUBMITTED')
                            st.session_state.submited = False
                            st.session_state.fac = facy
                            st.cache_data.clear()
                            st.cache_resource.clear()
                            st.session_state.submited =False
                            st.session_state.df = None
                            st.session_state.reader =False#
                            time.sleep(1)
                            st.rerun()
        
    if st.session_state.reader:# and st.session_state.df:
                    @st.cache_data
                    def lastqtr():
                        dat = ciraa.copy()
                        dat = dat[['ART', 'RD']].copy()
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE'})
                        return dat
                    @st.cache_data
                    def lastqt3():
                        dat = cirab.copy()
                        dat = dat[['ART', 'RD']].copy()
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE'})
                        return dat
                    @st.cache_data
                    def lost():
                        dat = df2wks.copy()
                        dat = dat[['ART', 'RD']].copy()
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE'})
                        return dat
                    @st.cache_data
                    def transfer():
                        dat = dft.copy()
                        dat = dat[['ART', 'RD', 'TO']]
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE', 'TO':'TRANSFER OUT DATE'})
                        return dat
                    @st.cache_data
                    def deceased():
                        dat = died.copy()
                        dat = dat[['ART', 'RD', 'DD']].copy()
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE', 'DD':'DEATH DATE'})
                        return dat
                    @st.cache_data
                    def viral():
                        dat = totalvl.copy()
                        dat = dat[['ART', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    ####ONE YEAR
                    @st.cache_data
                    def yearto():
                        dat = newto.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearlost():
                        dat = lostn.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.', 'AS':'ART START DATE','RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearvl():
                        dat = NVLa.copy()
                        dat = dat[['ART', 'AS','RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    ####6 MONTHS
                    @st.cache_data
                    def yearto6():
                        dat = newto6.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearlost6():
                        dat = lostn6.copy()
                        dat = dat[['ART', 'AS','RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearvl6():
                        dat = NVLa6.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.', 'AS':'ART START DATE','RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    ####3 MONTHS
                    @st.cache_data
                    def yearto3():
                        dat = newto3.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearlost3():
                        dat = lostn3.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearto3():
                        dat = newto3.copy()
                        dat = dat[['ART', 'AS','RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearlost3():
                        dat = lostn3.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearto1():
                        dat = newto1.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat
                    @st.cache_data
                    def yearlost1():
                        dat = lostn1.copy()
                        dat = dat[['ART','AS', 'RD', 'VD']]
                        dat = dat.rename(columns ={'ART':'ART NO.','AS':'ART START DATE', 'RD':'RETURN DATE', 'VD':'VIRAL LOAD DATE'})
                        return dat                  
                    preva = dfx[dfx['FACILITY']==facility] 
                    prev = preva['Q1'].sum()
                    fact = pot-prev
                    if fact < 2:
                            st.warning('THE POTENTIAL TX CURR IS LESS THAN THE Q1 CURR, WHICH MEANS AN ERROR WITH THIS EXTRACT')
                            time.sleep(1)
                            st.info('SHARE THIS WITH YOUR M AND E, TL OR TWG FOR MANUAL FILTERING')
                            err = 'ER'
                            time.sleep(3)
                    else: 
                            err = 'GD'
                    prev = int(prev)
                    part = [cluster,district,facility,week,wk,prev] #FIXED PART
                    part2 = [cluster, district, facility, week, wk]
                    list8 = part2 +list8
                    td = dt.date.today()
                    td = str(td)
                    list7 = [pot, txcurr, curto, deadcur, lostq, hasvl, td]
                    bands =['<01','01 to 09','10 to 19','20-29','30-39', '40-49','50+']
                
                    #cactive
                    #clost
                    ciralost = []
                    for band in bands:
                        clost['BAND'] = clost['BAND'].astype(str)
                        ct = clost[clost['BAND'] == band].copy()
                        ct['CIRAL'] =  ct['CIRAL'].astype(str)
                        dfk = ct[ct['CIRAL']=='<3 MTHS']
                        a = dfk.shape[0]
                        ciralost.append(a)
                        dfl = ct[ct['CIRAL']=='3-5 MTHS']
                        b = dfl.shape[0]
                        ciralost.append(b)
                        dfm = ct[ct['CIRAL']=='6 MTHS+']
                        c = dfm.shape[0]
                        ciralost.append(c)
                        
                    ciraactive =[]    
                    for band in bands:
                        cactive['BAND'] = cactive['BAND'].astype(str)
                        ct = cactive[cactive['BAND'] == band].copy()
                        ct['CIRAA'] =  ct['CIRAA'].astype(str)
                        dfk = ct[ct['CIRAA']=='<3 MTHS']
                        a = dfk.shape[0]
                        ciraactive.append(a)
                        dfl = ct[ct['CIRAA']=='3-5 MTHS']
                        b = dfl.shape[0]
                        ciraactive.append(b)
                        dfm = ct[ct['CIRAA']=='6 MTHS+']
                        c = dfm.shape[0]
                        ciraactive.append(c)
                    #pppp = cactive.copy()
                    row5 = part + ciralost + ciraactive
                    ell = list(err)
                  
                    # row1 = part + list1 + ell
                
                    # row2 = part + list2
                
                    # row3 = part + list3
                
                    # row4 = part + list4
                    row7 = part +list7
                        
                    secrets = st.secrets["connections"]["gsheets"]
                    
                        # Prepare the credentials dictionary
                    credentials_info = {
                            "type": secrets["type"],
                            "project_id": secrets["project_id"],
                            "private_key_id": secrets["private_key_id"],
                            "private_key": secrets["private_key"],
                            "client_email": secrets["client_email"],
                            "client_id": secrets["client_id"],
                            "auth_uri": secrets["auth_uri"],
                            "token_uri": secrets["token_uri"],
                            "auth_provider_x509_cert_url": secrets["auth_provider_x509_cert_url"],
                            "client_x509_cert_url": secrets["client_x509_cert_url"]
                        }
                            
                    try:
                        # Define the scopes needed for your application
                        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                                "https://www.googleapis.com/auth/drive"]
                        
                         
                        credentials = Credentials.from_service_account_info(credentials_info, scopes=scopes)
                            
                            # Authorize and access Google Sheets
                        client = gspread.authorize(credentials)
                            
                            # Open the Google Sheet by URL
                        spreadsheetu = "https://docs.google.com/spreadsheets/d/1twNlv9MNQWWsM73_dA19juHkp_Hua-k-fJA1qNVwQl0"
                        spreadsheet = client.open_by_url(spreadsheetu)
                    except Exception as e:
                            # Log the error message
                        st.write(f"CHECK: {e}")
                        st.write(traceback.format_exc())
                        st.write("COULDN'T CONNECT TO GOOGLE SHEET, TRY AGAIN")
                        st.stop()
                    ns = pd.read_csv('ALLNS.csv')
                    ns = ns[ns['DISTRICT']==district].copy()
                    ns = ns[ns['facility']==facility].copy()
                    ns = ns.rename(columns ={'ART': 'A'})
                    ns['A'] = pd.to_numeric(ns['A'], errors='coerce')
                    nsps['A'] = pd.to_numeric(nsps['A'], errors='coerce')
                    allns = pd.merge(ns, nsps, on ='A', how='left')
                    allns = allns.rename(columns ={'A': 'ARTN'})
                    allns['CLUSTER'] = np.nan
                    allns['CLUSTER'] = allns['CLUSTER'].fillna(cluster)
                    allns = allns[['CLUSTER','DISTRICT', 'facility','ARTN','ART','result_numeric', 'date_collected', 'AG','RD','LD', 'VD', 'TO','DD','RWEEK1','Ryear', 'Rmonth', 'Rday', 'Vyear', 'Vmonth', 'Vday', 'Lyear','Lmonth', 'Lday']]
                    allns = allns.rename(columns = {'RWEEK1': 'RWEEK'})
                    #LINE LISTS         
                    line[['Ryear', 'Rmonth', 'Rday']] = line[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                    line = line[((line['Ryear'] == 2025) & (line['Rmonth'].isin([1,2,3])))].copy()
                    tpt = line.copy()
                    cx = line.copy()
                    vl = line.copy()
                    pmtct = line.copy()
                    tpta  = tpt[tpt['TPT'].notna()].copy()
                    tptb  = tpt[tpt['TPT'].isnull()].copy()
                    tpta['TPT'] = tpta['TPT'].astype(str)
                    tpta = tpta[tpta['TPT']=='Never'].copy()
                    tpt = pd.concat([tpta, tptb])
                    month = dt.date.today().strftime('%m')
                    mon = int(month)
                    tpt[['Ayear', 'Amonth']] = tpt[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    tpta = tpt[((tpt['Ayear'] ==2024) & (tpt['Amonth'].isin([10,11,12])))].copy()
                    tptb = tpt[((tpt['Ayear'] <2024)| ((tpt['Ayear'] ==2024) & (tpt['Amonth']<10)))].copy() #NEXT Q ALL 2024 WILL BE ELIGIBLE
                    tpta[['Ayear', 'Rmonth']] = tpta[['Ayear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                    tpta['CHECK'] = tpt['Amonth']- tpt['Rmonth'].copy()
                    tpta['CHECK'] = pd.to_numeric(tpta['CHECK'], errors = 'coerce')
                    tpta = tpta[tpta['CHECK']<10].copy()
                    tpt = pd.concat([tpta, tptb])
                    #likely Vs unlikely
                    tpt[['Ayear', 'Amonth']] = tpt[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    tpta = tpt[((tpt['Ayear']<2024) | ((tpt['Ayear']==2024) & (tpt['Amonth'] <4)))].copy()
                    tptb = tpt[((tpt['Ayear']==2024) & (tpt['Amonth'] >3))].copy()
                    tpta['TPT STATUS'] = 'UNLIKELY'
                    tptb['TPT STATUS'] = 'LIKELY'
                    tpt = pd.concat([tpta, tptb])
                    tpt['Rmonth'] = pd.to_numeric(tpt['Rmonth'], errors = 'coerce')
                    jantpt = tpt[tpt['Rmonth']==1].shape[0]
                    febtpt = tpt[tpt['Rmonth']==2].shape[0]
                    martpt = tpt[tpt['Rmonth']==3].shape[0]
                    tpt = tpt[['A', 'TPT STATUS']] # GET RD,AS,RDAY,RMONTH, AFTER MERGING

                     #CERVICAL CANCER
                    cx['GD'] = cx['GD'].astype(str)
                    cx['GD'] = cx['GD'].str.replace('Female', 'F', regex=False)
                    cx['GD'] = cx['GD'].str.replace('FEMALE', 'F', regex=False)
                    cx = cx[cx['GD']=='F'].copy()
                    cx['AG'] = pd.to_numeric(cx['AG'], errors='coerce')
                    cx = cx[((cx['AG'] > 24) & (cx['AG'] < 50))].copy()
                    cxa = cx[cx['CX'].isnull()].copy()
                    cxb = cx[cx['CX'].notna()].copy()
                    cxb['CX'] = cxb['CX'].astype(str)
                    cxb = cxb[cxb['CX']== 'NOT ELIGIBLE'].copy()
                    cx = pd.concat([cxa,cxb])
                    cx['Rmonth'] = pd.to_numeric(cx['Rmonth'], errors = 'coerce')
                    jancx = cx[cx['Rmonth']==1].shape[0]
                    febcx = cx[cx['Rmonth']==2].shape[0]
                    marcx = cx[cx['Rmonth']==3].shape[0]
                    cx['CX STATUS'] = 'SCREEN'
                    cx = cx[['A', 'CX STATUS']].copy()
        
                    ###VL LINELIST
                    vl[['Ayear', 'Amonth']] = vl[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    vl = vl[((vl['Ayear']<2024) | ((vl['Ayear']==2024) & (vl['Amonth'] <10)))].copy() 
                    vl[['Ayear', 'Amonth']] = vl[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    vla = vl[((vl['Ayear']<2024) | ((vl['Ayear']==2024) & (vl['Amonth'] <7)))].copy() 
                    vlb = vl[((vl['Ayear']==2024) & (vl['Amonth'].isin([7,8,9])))].copy() 
                    vlb[['Amonth', 'Rmonth']] = vlb[['Amonth', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                    vlb['CHECK'] = vlb['Amonth'] - vlb['Rmonth'] 
                    vlb['CHECK'] = pd.to_numeric(vlb['CHECK'], errors = 'coerce')
                    vlb = vlb[vlb['CHECK']<7].copy()
                    vl = pd.concat([vla, vlb])
                    vl[['Vyear', 'Vmonth']] = vl[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                    vla = vl[((vl['Vyear'] < 2024) | ((vl['Vyear']==2024) & (vl['Vmonth'] <4)))].copy()
                    vla['VL STATUS'] = 'DUE'
                    vlx = vl[((vl['Vyear'] == 2024) & (vl['Vmonth'].isin([4,5])))].copy()
                    vl[['Vmonth', 'Rmonth']] = vl[['Vmonth', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                    vlb = vlx[(vlx['Vmonth']==4) &  (vlx['Rmonth'].isin([1,2]))].copy()
                    vlc = vlx[ ((vlx['Vmonth']==5) &  (vlx['Rmonth']==3))].copy()
                    vlc['TWOm'] = 'DUE'
                    vlb['TWOm'] = 'DUE'
                    vl = pd.concat([vla, vlb, vlc])
                    vl['Rmonth'] = pd.to_numeric(vl['Rmonth'], errors = 'coerce')
                    janvl = vl[vl['Rmonth']==1].shape[0] 
                    febvl = vl[vl['Rmonth']==2].shape[0]
                    marvl = vl[vl['Rmonth']==3].shape[0]
                    vl = vl[['A', 'VL STATUS', 'TWOm']].copy()
     

                    #MERGING TPT AND VL LISTS
                    vl['A'] = pd.to_numeric(vl['A'], errors ='coerce')
                    tpt['A'] = pd.to_numeric(tpt['A'], errors ='coerce')
                    linea = pd.merge(vl, tpt , on = 'A', how='outer')

                     #MERGING THE NEW LIST ABOVE AND THE CX LIST
                    linea['A'] = pd.to_numeric(linea['A'], errors ='coerce')
                    cx['A'] = pd.to_numeric(cx['A'], errors ='coerce')
                    lineb = pd.merge(linea, cx, on = 'A', how='outer')

                    #PMTCT VL LIST
                    pmtct['PRG'] = pmtct['PT'].astype(str)
                    pmtct['PRG'] = pmtct['PRG'].str.replace('Yes', 'YES')
                    pmtct['PRG'] = pmtct['PRG'].str.replace('Breast feeding', 'YES')
                    pmtct = pmtct[pmtct['PRG']=='YES'].copy()
                    pmtct[['Vyear', 'Vmonth']] = pmtct[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                    pmtct = pmtct[pmtct['Vyear']<2025].copy()
                    pmtct['PVL'] = 'DUE'
                    pmtct = pmtct[['A', 'PVL']].copy()

                    #MERGING THE LIST ABOVE AND THE PMTCT LIST
                    lineb['A'] = pd.to_numeric(lineb['A'], errors ='coerce')
                    pmtct['A'] = pd.to_numeric(pmtct['A'], errors ='coerce')
                    linec = pd.merge(lineb, pmtct, on = 'A', how = 'outer')
                    
                    #MERGING THE ABOVE LIST WITH THE ORIGINAL LIS
                    linec['A'] = pd.to_numeric(linec['A'], errors ='coerce')
                    line['A'] = pd.to_numeric(line['A'], errors ='coerce')
                    line = pd.merge(linec, line, on = 'A', how = 'left')
                    line['CLUSTER'] = cluster
                    line['DISTRICT'] = district
                    line['FACILITY']= facility
                    line = line[['CLUSTER', 'DISTRICT', 'FACILITY', 'A','AG','GD', 'AS', 'RD', 'VD', 'Ryear', 'Rmonth', 'Rday', 'RWEEK1','VL STATUS', 'TWOm', 'TPT', 'TPT STATUS', 'CX', 'CX STATUS', 'PT' , 'PVL']].copy()
                    line = line.rename(columns={'RWEEK1': 'RWEEK'})
                    missed[['Lyear', 'Lmonth']] = missed[['Lyear', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                    ma = dt.date.today()
                    me  = ma.strftime('%m')
                    mot = int(me) 
                    #missed = missed[((missed['Lyear']==2025) & (missed['Lmonth']<=mot))].copy()
                    missed = missed[((missed['Lyear']==2024) & (missed['Lmonth']==11))].copy()

                    #VL MISSED

                    tptmis = missed.copy()
                    cxmis = missed.copy()
                    vlmis = missed.copy()
                 
                    tptamis  = tptmis[tptmis['TPT'].notna()].copy()
                    tptbmis  = tptmis[tptmis['TPT'].isnull()].copy()
                    tptamis['TPT'] = tptamis['TPT'].astype(str)
                    tptamis = tptamis[tptamis['TPT']=='Never'].copy()
                    tptmis = pd.concat([tptamis, tptbmis])
                    month = dt.date.today().strftime('%m')
                    mon = int(month)
                    tptmis[['Ayear', 'Amonth']] = tptmis[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    tptamis = tptmis[((tptmis['Ayear'] ==2024) & (tptmis['Amonth'].isin([10,11,12])))].copy()
                    tptbmis = tptmis[((tptmis['Ayear'] <2024)| ((tptmis['Ayear'] ==2024) & (tptmis['Amonth']<10)))].copy() #NEXT Q ALL 2024 WILL BE ELIGIBLE
                    tptamis[['Ayear', 'Lmonth']] = tptamis[['Ayear', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                    tptamis['CHECK'] = tptmis['Amonth']- tptmis['Lmonth'].copy()
                    tptamis['CHECK'] = pd.to_numeric(tptamis['CHECK'], errors = 'coerce')
                    tptamis = tptamis[tptamis['CHECK']<10].copy()
                    tptmis = pd.concat([tptamis, tptbmis])
                    tptmis['TPT STATUS'] = 'NOT INITIATED'
                    tptmis = tptmis[['A','TPT', 'TPT STATUS']]
                    notpt = tptmis.shape[0]
      
                     #CERVICAL CANCER
                    cxmis['GD'] = cxmis['GD'].astype(str)
                    cxmis['GD'] = cxmis['GD'].str.replace('Female', 'F', regex=False)
                    cxmis['GD'] = cxmis['GD'].str.replace('FEMALE', 'F', regex=False)
                    cxmis = cxmis[cxmis['GD']=='F'].copy()
                    cxmis['AG'] = pd.to_numeric(cxmis['AG'], errors='coerce')
                    cxmis = cxmis[((cxmis['AG'] > 24) & (cxmis['AG'] < 50))].copy()
                    cxamis = cxmis[cxmis['CX'].isnull()].copy()
                    cxbmis = cxmis[cxmis['CX'].notna()].copy()
                    cxbmis['CX'] = cxbmis['CX'].astype(str)
                    cxbmis = cxbmis[cxbmis['CX']== 'NOT ELIGIBLE'].copy()
                    cxmis = pd.concat([cxamis,cxbmis])
                    cxmis['CX STATUS'] = 'SCREEN'
                    cxmis = cxmis[['A', 'CX STATUS']].copy()
                    notscreened = cxmis.shape[0]
        
                    ###VL LINELIST
                    vlmis[['Ayear', 'Amonth']] = vlmis[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    vlmis = vlmis[((vlmis['Ayear']<2024) | ((vlmis['Ayear']==2024) & (vlmis['Amonth'] <10)))].copy() 
                    vlmis[['Ayear', 'Amonth']] = vlmis[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                    vlamis = vlmis[((vlmis['Ayear']<2024) | ((vlmis['Ayear']==2024) & (vlmis['Amonth'] <7)))].copy() 
                    vlbmis = vlmis[((vlmis['Ayear']==2024) & (vlmis['Amonth'].isin([7,8,9])))].copy() 
                    vlbmis[['Amonth', 'Rmonth']] = vlbmis[['Amonth', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                    vlbmis['CHECK'] = vlbmis['Amonth'] - vlbmis['Lmonth'] 
                    vlbmis['CHECK'] = pd.to_numeric(vlbmis['CHECK'], errors = 'coerce')
                    vlbmis = vlbmis[vlbmis['CHECK']<7].copy()
                    vlmis = pd.concat([vlamis, vlbmis])
                    vlmis[['Vyear', 'Vmonth']] = vlmis[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                    vlamis = vlmis[((vlmis['Vyear'] < 2024) | ((vlmis['Vyear']==2024) & (vlmis['Vmonth'] <4)))].copy()
                    vlamis['VL STATUS'] = 'DUE'
                    vlxmis = vlmis[((vlmis['Vyear'] == 2024) & (vlmis['Vmonth'].isin([4,5])))].copy()
                    vlmis[['Vmonth', 'Lmonth']] = vlmis[['Vmonth', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                    vlbmis = vlxmis[(vlxmis['Vmonth']==4) &  (vlxmis['Rmonth'].isin([1,2]))].copy()
                    vlcmis = vlxmis[ ((vlxmis['Vmonth']==5) &  (vlxmis['Rmonth']==3))].copy()
                    vlcmis['TWOm'] = 'DUE'
                    vlbmis['TWOm'] = 'DUE'
                    vlmis = pd.concat([vlamis, vlbmis, vlcmis])
                    vlmis = vlmis[['A', 'VL STATUS', 'TWOm']].copy()
                    notbled = vlmis.shape[0]

                    #MERGING TPT AND VL LISTS
                    vlmis['A'] = pd.to_numeric(vlmis['A'], errors ='coerce')
                    tptmis['A'] = pd.to_numeric(tptmis['A'], errors ='coerce')
                    lineamis = pd.merge(vlmis, tptmis , on = 'A', how='outer')

                     #MERGING THE NEW LIST ABOVE AND THE CX LIST
                    lineamis['A'] = pd.to_numeric(lineamis['A'], errors ='coerce')
                    cxmis['A'] = pd.to_numeric(cxmis['A'], errors ='coerce')
                    linebmis = pd.merge(lineamis, cxmis, on = 'A', how='outer')
 
                    missed = missed[['A', 'AG', 'GD', 'RD','LD']].copy() 
                    missed['A'] = pd.to_numeric(missed['A'], errors ='coerce')
                    linebmis['A'] = pd.to_numeric(linebmis['A'], errors ='coerce')
                    missed = pd.merge(linebmis, missed, on = 'A', how = 'left')
                    mmm = missed.copy()
                    
                    @st.cache_data
                    def missedlists():
                        dat = mmm.copy()
                        dat = dat.rename(columns={'LD': 'LAST ENCOUNTER', 'GD':'GENDER','AG':'AGE', 'RD':'RETURN DATE', 'A':'ART No.'})
                        dat = dat[['ART No.','AGE', 'RETURN DATE','GENDER',  'LAST ENCOUNTER', 'VL STATUS', 'TWOm', 'TPT', 'TPT STATUS', 'CX STATUS']].copy()
                        return dat

                    #SUMMARY LINELIST
                    col1,col2,col3 = st.columns([1,2,1])
                    with col3:
                         submit = st.button('Submit') 

                    linelists = [cluster, district, facility, jancx, janvl,jantpt, febcx, febvl, febtpt, marcx, marvl, martpt, notbled, notpt, notscreened, wk]
        
                    if submit:
                            # conn = st.connection('gsheets', type=GSheetsConnection)
                            # exist = conn.read(worksheet= 'ALLNS', usecols=list(range(24)),ttl=5)
                            # existing= exist.dropna(how='all')
                            # checkf = existing['facility'].unique()
                            # if facility in checkf:
                            #     pass
                            # else:
                            #     updated = pd.concat([existing, allns], ignore_index =True)
                            #     conn.update(worksheet = 'ALLNS', data = updated) 
                            try:
                                # conn = st.connection('gsheets', type=GSheetsConnection)
                                # exist = conn.read(worksheet= 'LINELISTS', usecols=list(range(22)),ttl=5)
                                # exist2 = conn.read(worksheet= 'SUMM', usecols=list(range(16)),ttl=5)
                                # dfcheck = exist2.dropna(how='all')
                                # dfcheck['WEEK'] = pd.to_numeric(dfcheck['WEEK'], errors='coerce')
                                # dfcheck = dfcheck[dfcheck['WEEK']==wk].copy()
                                # facitiz = dfcheck['FACILITY'].unique()
                                # if facility in facitiz:
                                #     pass
                                # else:
                                #     dfex = exist.dropna(how='all')
                                #     dfex = dfex[dfex['FACILITY']!=facility].copy()
                                #     line['RWEEK'] = pd.to_numeric(line['RWEEK'], errors= 'coerce')
                                #     wkapp = wk +1
                                #     line = line[((line['RWEEK']==wk) | (line['RWEEK']== wkapp))].copy()
                                #     dfline = pd.concat([dfex, line])
                                #     conn.update(worksheet = 'LINELISTS', data = dfline)
                                #     sheet6 = spreadsheet.worksheet("SUMM")
                                    
                                #     sheet6.append_row(linelists, value_input_option='RAW')
                                
                                # sheet1 = spreadsheet.worksheet("TX")
                                # #st.write(row1)
                                # sheet1.append_row(row1, value_input_option='RAW')
                                    
                                # sheet2 = spreadsheet.worksheet("VL")
                                # sheet2.append_row(row2, value_input_option='RAW')
                                    
                                # sheet3 = spreadsheet.worksheet("YEARS")
                                # sheet4 = spreadsheet.worksheet("THREEO")
                                # sheet5 = spreadsheet.worksheet("CIRA")
                                
                                # sheet3.append_row(row3, value_input_option='RAW')
                                # sheet4.append_row(row4, value_input_option='RAW')
                                # sheet5.append_row(row5, value_input_option='RAW')
                                sheet7 = spreadsheet.worksheet("Q4")
                                sheet7.append_row(row7, value_input_option='RAW')
                                sheet9 = spreadsheet.worksheet("ONEYR")
                                sheet9.append_row(list8, value_input_option='RAW')
                                st.session_state.submited = True
                            except Exception as e:
                                # Print the error message
                                st.write(f"ERROR: {e}")
                                st.stop()  # Stop the Streamlit app here to let the user manually retry     
                    else:
                            st.write('FIRST SUBMIT TO SEE LINELISTS AND SUMMARY') 
                            st.markdown(f'**YOU HAVE SELECTED {district} AS THE DISTRICT AND {facility} AS THE FACILITY**')
                            st.write('BE SURE OF THE ABOVE SELECTIONS BEFORE SUBMITTING')                     
                    
                    if not st.session_state.submited:
                        st.stop()  
                    if str(facy) != str(st.session_state.fac):
                        st.stop()
                    if st.session_state.submited:
                            st.success('**SUBMITTED, To upload another excel, first refresh this page, or open the link afresh**')
                            st.divider()
                            if facility == 'Kifampa HC III':
                                pass
                            elif pot < prev:
                                st.info('**SOMETHING IS WRONG WITH THIS EXTRACT, SEND TO YOUR M AND E TO CHECK, POTENTIAL TX CURR IS LESS THAN Q1 CURR**')
                                st.stop()
                            else:
                                pass 
                            st.write(f'**Q1 CURR: {prev}, CURRENT CURR {txcurr}, VL COVERAGE {vlcov} %**')
                            if curr < prev:
                                st.write(f'**Banange, 😢 you have dropped this TX CURR BY {prev-txcurr}** 😢😢😢' )
                            elif curr == prev:
                                st.info('**THANK YOU, YOU HAVE ACHIEVED YOUR Q1 TX CURR 👏👏👏👏**')
                                st.balloons()
                                time.sleep(2)
                                st.balloons()
                            elif curr > prev:
                                st.info(f'**CONGRATULATIONS, YOU HAVE EXCEEDED YOUR Q1 CURR BY {curr-prev} 👏👏👏👏**')
                                st.balloons()
                                time.sleep(2)
                                st.balloons()
                                time.sleep(2)
                                st.balloons()
                            st.write(f"<h6><b>DOWNLOAD LINELISTS FROM HERE</b></h6>", unsafe_allow_html=True)
                            cola, colb, colc = st.columns(3)
                            with cola:
                                    if lostq ==0:
                                        st.write('**NO TXML**')
                                    else:
                                       dat = currlost.copy() 
                                       #dat = tttt.copy()
                                       csv_data = dat.to_csv(index=False)
                                       tot = dat.shape[0]
                                       st.write(f'**YOUR TXML IS {tot}**')
                                       st.download_button(
                                                    label="Q2 TXML",
                                                    data=csv_data,
                                                    file_name=f"{facility} TXML.csv",
                                                    mime="text/csv")
                            with colb:
                                    if lacks.shape[0] ==0:
                                        st.write('**NO VL LINELIST**')
                                    else:
                                       dat = lacks.copy() 
                                       csv_data = dat.to_csv(index=False)
                                       tot = dat.shape[0]
                                       st.write(f'**{tot} CLIENTS ARE DUE FOR VL**')
                                       st.download_button(
                                                    label="VL LINELISTS",
                                                    data=csv_data,
                                                    file_name=f"{facility} VL_LINELIST.csv",
                                                    mime="text/csv")

                        
                        #             if two==0:
                        #                 st.write('**NO MISSED APPOINTMENTS**')
                        #             else:
                        #                 dat = lost()
                        #                 #dat = yyy.copy()
                        #                 csv_data = dat.to_csv(index=False)
                        #                 tot = dat.shape[0]
                        #                 st.write(f'**{tot} CLIENTS HAVE MISSED**')
                        #                 st.download_button(
                        #                             label="MISSED APPOINTMENTS",
                        #                             data=csv_data,
                        #                             file_name=f"{facility} MISSED.csv",
                        #                             mime="text/csv")
                        #     with colb:
                        #             if dead ==0:
                        #                 st.write('**NO DEAD CLIENTS**')
                        #             else:
                        #                 dat = deceased()
                        #                 csv_data = dat.to_csv(index=False)
                        #                 tot = dat.shape[0]
                        #                 st.write(f'**{tot} DEAD THIS Qtr**')
                        #                 st.download_button(
                        #                                 label=" DEAD",
                        #                                 data=csv_data,
                        #                                 file_name=f" {facility} DEAD.csv",
                        #                                 mime="text/csv")
                        #     with colc:
                        #             if true == 0:
                        #                 st.markdown('**NO TRANSFER OUTs**')
                        #             else:
                        #                 dat = transfer()
                        #                 csv_data = dat.to_csv(index=False)
                        #                 tot = dat.shape[0]
                        #                 st.write(f'**{tot} TOs THIS Qtr**')
                        #                 st.download_button(
                        #                             label=" TRANSFER OUTS",
                        #                             data=csv_data,
                        #                             file_name=f" {facility} TOS.csv",
                        #                             mime="text/csv")
                        # ######################################VL SECTION
                        #     st.divider()
                        #     st.markdown("**LAST QUARTER'S TXML AND VIRAL LOAD LINE LIST**")
                        #     cola, colb,colc = st.columns(3)
                        #     with cola:
                        #         dat = lastqtr()
                        #         csv_data = dat.to_csv(index=False)
                        #         tot = dat.shape[0]
                        #         st.write(f'**{tot} NOT RETURNED FROM Q4**')
                        #         st.download_button(
                        #                     label="TXML FOR Q4",
                        #                     data=csv_data,
                        #                     file_name=f"{facility} Q4_TXML.csv",
                        #                     mime="text/csv")
                        #     with colb:
                        #         dat = lastqt3()
                        #         csv_data = dat.to_csv(index=False)
                        #         tot = dat.shape[0]
                        #         st.write(f'**{tot} NOT RETURNED FROM Q3**')
                        #         st.download_button(
                        #                     label="TXML FOR Q3",
                        #                     data=csv_data,
                        #                     file_name=f"{facility} Q3_TXML.csv",
                        #                     mime="text/csv")
                        #     with colc:
                        #         dat = viral()
                        #         csv_data = dat.to_csv(index=False)
                        #         tot = dat.shape[0]
                        #         st.write(f'**{tot} DUE FOR VL**')
                        #         st.download_button(
                        #                     label="CURRENT VL LINELIST",
                        #                     data=csv_data,
                        #                     file_name=f"{facility} VL.csv",
                        #                     mime="text/csv")              
                            
                        # #     #########################################################################################################################################################
                        # ###ONE YEAR LINE LISTS
                        #     if st.session_state.submited: 
                        #         st.divider()
                        #         st.write(f"<h6><b>ONE YEAR COHORT LINELISTS </b></h6>", unsafe_allow_html=True)
                        #         cola, colb, colc = st.columns(3)
                        #         with cola:
                        #                 if newlost==0:
                        #                     st.write('**NO 1 YR IIT**')
                        #                 else:
                        #                     dat = yearlost()
                        #                     csv_data = dat.to_csv(index=False)
                        #                     tot = dat.shape[0]
                        #                     st.write(f'**{tot} LTFU IN THE 1 YR COHORT**')
                        #                     st.download_button(key='a',
                        #                                 label="ONE YR IIT",
                        #                                 data=csv_data,
                        #                                 file_name=f"{facility} 1YR_IIT.csv",
                        #                                 mime="text/csv")
                        #         with colb:
                        #                 if outnew==0:
                        #                     st.markdown('**NO 1 YR TOs**')
                        #                 else:
                        #                     dat = yearto()
                        #                     csv_data = dat.to_csv(index=False)
                        #                     tot = dat.shape[0]
                        #                     st.write(f'**{tot} TOs**')
                        #                     st.download_button(key='b',
                        #                                 label=" 1 YR T.OUTS",
                        #                                 data=csv_data,
                        #                                 file_name=f" {facility} TO_1YR.csv",
                        #                                 mime="text/csv")
                        #         with colc:
                        #             if nvla ==0:
                        #                 st.write('**NO ONE YEAR VL LIST**')
                        #             else:
                        #                 dat = yearvl()
                        #                 csv_data = dat.to_csv(index=False)
                        #                 tot = dat.shape[0]
                        #                 st.write(f'**{tot} HAVE NOT BEEN BLED**')
                        #                 st.download_button(key='c',
                        #                             label="1 YR VL LINELIST",
                        #                             data=csv_data,
                        #                             file_name=f"{facility} VL_1YR.csv",
                        #                             mime="text/csv")
                                
                        #     ###SIX YEAR LINE LISTS
                        #         st.divider() 
                        #         st.write(f"<h6><b>SIX MONTHS COHORT LINELISTS </b></h6>", unsafe_allow_html=True)
                        #         cola, colb, colc = st.columns(3)
                        #         with cola:
                        #                 if newlost6==0:
                        #                     st.write('**NO 6 MTHS IIT**')
                        #                 else:
                        #                     dat = yearlost6()
                        #                     csv_data = dat.to_csv(index=False)
                        #                     tot = dat.shape[0]
                        #                     st.write(f'**{tot} LTFUS**')
                        #                     st.download_button(key='d',
                        #                                 label="SIX MTHS IIT",
                        #                                 data=csv_data,
                        #                                 file_name=f"{facility} IIT_6.csv",
                        #                                 mime="text/csv")
                        #         with colb:
                        #                 if outnew6==0:
                        #                     st.markdown('**NO 6 MTHS TOs**')
                        #                 else:
                        #                     dat = yearto6()
                        #                     csv_data = dat.to_csv(index=False)
                        #                     tot = dat.shape[0]
                        #                     st.write(f'**{tot} TOs**')
                        #                     st.download_button(key='e',
                        #                                 label=" 6 MTHS T.OUTS",
                        #                                 data=csv_data,
                        #                                 file_name=f" {facility} TO_1YR.csv",
                        #                                 mime="text/csv")
                        #         with colc:
                        #             if nvla6 ==0:
                        #                 st.markdown('**NO 6 MTHS VL LIST**')
                        #             else:
                        #                 dat = yearvl6()
                        #                 csv_data = dat.to_csv(index=False)
                        #                 tot = dat.shape[0]
                        #                 st.write(f'**{tot} DUE FOR FIRST VL**')
                        #                 st.download_button(key='f',
                        #                             label="6 MTHS VL",
                        #                             data=csv_data,
                        #                             file_name=f"{facility} VL6.csv",
                        #                             mime="text/csv")
                                                            
                            # ###THREE MTHS LINE LISTS
                            #     st.divider()
                            #     st.write(f"<h6><b>THREE MONTHS COHORT LINELISTS </b></h6>", unsafe_allow_html=True)
                            #     cola, colb = st.columns(2)
                            #     with cola:
                            #             if newlost3==0:
                            #                 st.write('**NO 3 MTHS IIT**')
                            #             else:
                            #                 dat = yearlost3()
                            #                 csv_data = dat.to_csv(index=False)
                            #                 st.download_button(key='g',
                            #                             label="3 MTHS IIT",
                            #                             data=csv_data,
                            #                             file_name=f"{facility} IIT_3.csv",
                            #                             mime="text/csv")
                            #     with colb:
                            #             # if outnew3==0:
                            #             #     st.markdown('**NO 3 MTHS TOs**')
                            #             # else:
                            #                 dat = yearto3()
                            #                 dat = yyyu.copy()
                            #                 csv_data = dat.to_csv(index=False)
                            #                 st.download_button(key='h',
                            #                             label="3 MTHS T.OUTS",
                            #                             data=csv_data,
                            #                             file_name=f" {facility} TOs_3.csv",
                            #                             mime="text/csv")                    
                            
                            # ###THREE MTHS LINE LISTS   
                            #     st.divider()
                            #     st.write(f"<h6><b>TX NEW LINELISTS </b></h6>", unsafe_allow_html=True)
                            #     cola, colb = st.columns(2)
                            #     with cola:
                            #             if newlost1==0:
                            #                 st.write('**NO TX NEW IIT**')
                            #             else:
                            #                 dat = yearlost1()
                            #                 csv_data = dat.to_csv(index=False)
                            #                 st.download_button(key='j',
                            #                             label="TX NEW IIT",
                            #                             data=csv_data,
                            #                             file_name=f"{facility} IIT_NEW.csv",
                            #                             mime="text/csv")
                            #     with colb:
                            #         if outnew1==0:
                            #                 st.markdown('**NO TxNEW TOs**')
                            #         else:
                            #             dat = yearto1()
                            #             csv_data = dat.to_csv(index=False)
                            #             st.download_button(key='k',
                            #                         label="TXNEW T.OUTS",
                            #                         data=csv_data,
                            #                         file_name=f" {facility} TxNEW_TOs.csv",
                            #                         mime="text/csv") 
                            #     st.divider()
                            #     forth = forth.rename(columns = {'A': 'ART NO.','VD': 'VL DATE', 'RD': 'RETURN DATE', 'DD': 'DEATH DATE', 'TO': 'TRANSFER OUT DATE', 'AS': 'ART START DATE'})
                            #     cola,colb = st.columns([4,1])
                            #     with cola:
                            #             st.markdown('**MASTER LIST WITH ALL LINELISTS COMBINED**')
                            #             dat = forth.copy()
                            #             #dat = pppp.copy()
                            #             csv_data = dat.to_csv(index=False) 
                            #             st.download_button(
                            #                         label="MASTER_LIST",
                            #                         data=csv_data,
                            #                         file_name=f" {facility} MASTER_LIST.csv",
                            #                         mime="text/csv")
                                    
                            #     cola,colb = st.columns([4,1])
                            #     with cola:
                            #             st.markdown('**CLINETS THAT MISSED SERVICES**')
                            #             dat = missedlists()
                            #             dat = yyyuu.copy()
                            #             csv_data = dat.to_csv(index=False) 
                            #             st.download_button(
                            #                         label="MISSED OPPORTUNITIES",
                            #                         data=csv_data,
                            #                         file_name=f" {facility} MASTER_LIST.csv",
                            #                         mime="text/csv")
                
                            st.divider()
                                # st.write(FYA)
                                # st.write(FYB)
                                # st.write(FYA)
                            st.success('**WANT TO HELP US IMPROVE?**')
                            st.write('Are you getting different results when you filter the extract manually?, That is ok')
                            st.write('**The intention of this program is to get the same results as you would manually, so help us improve by sending any variation you get to the TWG**')
                            st.warning('Refer to the SOP section to see how this program arrives to the summaries and linelists you are seeing')
                            st.write('')
                            st.write('')
                            st.write('')
                            st.success('**CREATED BY Dr. LUMINSA DESIRE**')
                            st.info('**WITH CONTRIBUTION FROM EDISON KATUNGUKA, SIMON SEMAKULA AND CHRIS MUGARA, FOR THE TWG**')
pages = {
    "READER:": [
        st.Page(extract, title="EMR EXTRACT READER"),
    ],
    "VISUALISATION:":[
        st.Page("TXML_CAMPAIGN.py", title="TXML CAMPAIGN"),],
        # st.Page("VL_SECTION.py", title="LINELISTS")],
    "RESOURCES:": [
        st.Page("SOPs.py", title="SOPs"),
        st.Page("USER_MANUAL.py", title="USER MANUAL"),
    ],
}

pg = st.navigation(pages)
pg.run()
                                
    

